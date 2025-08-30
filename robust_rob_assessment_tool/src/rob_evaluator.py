import os
import glob
import pandas as pd
import logging
import re
import json
import openpyxl
from openpyxl.styles import PatternFill
import time
import pickle
from tqdm import tqdm
import warnings
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Optional, Any
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

from document_processor import DocumentProcessor
from llm_config import LLMConfig
from data_models import CoreAssessmentResponse, OptionalAssessmentResponse
from error_handler import ErrorHandler, ROBError, ErrorCategory, ErrorSeverity
from cost_analyzer import CostAnalyzer

warnings.filterwarnings('ignore')
logger = logging.getLogger(__name__)

class ROBEvaluator:
    def __init__(self, config: dict, batch_id: Optional[str] = None):
        """
        Initialize ROB Evaluator with optional batch processing support
        
        Args:
            config: Configuration dictionary
            batch_id: Optional batch identifier for parallel processing
        """
        self.batch_id = batch_id
        self.config = config
        
        # Parallel processing configuration
        self.parallel_enabled = config.get('parallel', {}).get('enabled', False)
        self.max_workers = config.get('parallel', {}).get('max_workers', 1)
        self.parallel_workers = config.get('parallel', {}).get('parallel_workers', self.max_workers)
        
        # Thread-safe locks for parallel processing
        self.results_lock = threading.Lock()
        self.checkpoint_lock = threading.Lock()
        self.cache_lock = threading.Lock()
        self.llm_configs = [
            LLMConfig(
                name=model_config["name"],
                api_key=model_config["api_key"],
                base_url=model_config["base_url"],
                model_name=model_config["model_name"],
                use_streaming=model_config.get("use_streaming", True)
            )
            for model_config in config["llm_models"]
        ]

        self.output_mode = config["processing"].get("llm_output_mode", "json")
        logger.info(f"LLM output mode set to: {self.output_mode.upper()}")
        
        self.eval_optional_items = config["processing"]["eval_optional_items"]
        self.max_text_length = config["processing"]["max_text_length"]
        self.checkpoint_file = config["paths"]["checkpoint_file"]
        self.start_index = config["processing"]["start_index"]
        
        # Batch-specific checkpoint file if batch_id is provided
        if self.batch_id:
            checkpoint_dir = os.path.dirname(self.checkpoint_file)
            checkpoint_name = f"checkpoint_batch_{self.batch_id}.pkl"
            self.checkpoint_file = os.path.join(checkpoint_dir, checkpoint_name)
        
        self.domain6_thresholds = config["domain6"]["thresholds"]
        self.domain6_default = config["domain6"]["default_assessment"]
        
        self.results = {config.name: {'core': [], 'optional': []} for config in self.llm_configs}
        self.documents_cache = {}
        self.processed_files = self._load_checkpoint()
        self.study_id_cache = {}
        
        # Batch processing state
        self.batch_errors = []
        self.batch_stats = {
            'total_documents': 0,
            'processed_documents': 0,
            'failed_documents': 0,
            'start_time': None,
            'end_time': None
        }
        
        # Initialize error handler
        error_config = {
            'retry_attempts': config.get('error_handling', {}).get('retry_attempts', 3),
            'base_delay': config.get('error_handling', {}).get('base_delay', 2),
            'max_delay': config.get('error_handling', {}).get('max_delay', 60)
        }
        self.error_handler = ErrorHandler(error_config)
        
        # Initialize cost analyzer
        pricing_config_path = config["paths"].get("llm_pricing_config", 
                                                 os.path.join(os.path.dirname(os.path.dirname(__file__)), 
                                                             "config", "llm_pricing.json"))
        session_id = f"rob_session_{self.batch_id}" if self.batch_id else f"rob_session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        try:
            self.cost_analyzer = CostAnalyzer(pricing_config_path, session_id)
            logger.info(f"Cost analyzer initialized with session ID: {session_id}")
        except Exception as e:
            logger.warning(f"Failed to initialize cost analyzer: {e}. Cost tracking will be disabled.")
            self.cost_analyzer = None

    def _load_checkpoint(self):
        """Load checkpoint data"""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'rb') as f:
                    checkpoint_data = pickle.load(f)
                logger.info(f"Checkpoint data loaded: {len(checkpoint_data)} file-model combinations processed")
                return checkpoint_data
            except Exception as e:
                logger.error(f"Failed to load checkpoint data: {e}")
                return set()
        return set()

    def _save_checkpoint(self, file_path: str, model_name: str):
        """Save checkpoint data (thread-safe)"""
        checkpoint_key = f"{model_name}:{file_path}"
        self.processed_files.add(checkpoint_key)
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(self.checkpoint_file), exist_ok=True)
            with open(self.checkpoint_file, 'wb') as f:
                pickle.dump(self.processed_files, f)
            logger.debug(f"Checkpoint data updated: {len(self.processed_files)} total file-model combinations")
        except Exception as e:
            logger.error(f"Failed to save checkpoint data: {e}")

    def process_folder(self, folder_path: str, output_path: str):
        """Process all documents in the folder with optional parallel processing"""
        # Search for files recursively, including subdirectories
        pdf_files = glob.glob(os.path.join(folder_path, "**", "*.pdf"), recursive=True)
        docx_files = glob.glob(os.path.join(folder_path, "**", "*.docx"), recursive=True)
        files = pdf_files + docx_files
        valid_files = [f for f in files if os.path.isfile(f)]
        
        logger.info(f"Found {len(valid_files)} files to process")
        
        if self.start_index > 0:
            if self.start_index >= len(valid_files):
                logger.warning(f"Start index {self.start_index} exceeds total files {len(valid_files)}, nothing to process")
                return
            valid_files = valid_files[self.start_index:]
            logger.info(f"Skipping first {self.start_index} files, starting at file {self.start_index+1}")
        
        os.makedirs(output_path, exist_ok=True)
        output_file = os.path.join(output_path, "rob_results.xlsx")
        
        # Check if parallel processing is enabled and we have multiple workers
        if self.parallel_enabled and self.parallel_workers > 1:
            logger.info(f"🚀 Using parallel processing with {self.parallel_workers} workers")
            self._process_folder_parallel(valid_files, output_path)
        else:
            logger.info("Using sequential processing")
            self._process_folder_sequential(valid_files, output_path)

        logger.info("Entire evaluation process completed!")

    def _process_folder_sequential(self, valid_files: List[str], output_path: str):
        """Process documents sequentially (original method)"""
        for idx, file_path in enumerate(tqdm(valid_files, desc="Processing documents")):
            self._process_single_document(file_path, idx, len(valid_files), output_path)

    def _process_folder_parallel(self, valid_files: List[str], output_path: str):
        """Process documents in parallel using ThreadPoolExecutor"""
        completed_count = 0
        total_files = len(valid_files)
        
        # Create a progress bar
        progress_bar = tqdm(total=total_files, desc="Processing documents (parallel)")
        
        # Use ThreadPoolExecutor for parallel processing
        with ThreadPoolExecutor(max_workers=self.parallel_workers) as executor:
            # Submit all tasks
            future_to_file = {
                executor.submit(self._process_single_document, file_path, idx, total_files, output_path): file_path
                for idx, file_path in enumerate(valid_files)
            }
            
            # Process completed tasks
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    future.result()  # This will raise any exception that occurred
                    completed_count += 1
                    progress_bar.update(1)
                    progress_bar.set_postfix({
                        'completed': completed_count,
                        'workers': self.parallel_workers
                    })
                except Exception as e:
                    logger.error(f"Error processing {file_path}: {e}")
                    progress_bar.update(1)
        
        progress_bar.close()
        logger.info(f"✅ Parallel processing completed: {completed_count}/{total_files} files processed successfully")

    def _process_single_document(self, file_path: str, idx: int, total_files: int, output_path: str):
        """Process a single document (thread-safe)"""
        file_name = os.path.basename(file_path)
        thread_id = threading.current_thread().name
        logger.info(f"[{thread_id}] Processing document [{self.start_index+idx+1}/{total_files+self.start_index}]: {file_name}")
        
        # Extract text
        text = DocumentProcessor.extract_text(file_path)
        
        if not text:
            logger.warning(f"[{thread_id}] Failed to extract text: {file_name}")
            return
        
        # Thread-safe study ID extraction/caching
        with self.cache_lock:
            if file_path not in self.study_id_cache:
                logger.info(f"[{thread_id}] Extracting study ID from document {file_name}")
                study_id = self.extract_study_id(text, file_name, self.llm_configs[0])
                self.study_id_cache[file_path] = study_id
                logger.info(f"[{thread_id}] Extracted study ID: {study_id}")
            else:
                study_id = self.study_id_cache[file_path]
                logger.info(f"[{thread_id}] Using cached study ID: {study_id}")
        
        logger.info(f"[{thread_id}] Total document length: {len(text)} characters, limit: {self.max_text_length} characters")
        
        content = {
            'text': text,
            'study_id': study_id,
            'file_name': file_name
        }
        
        # Evaluate the current document with each model
        for model_idx, llm_config in enumerate(self.llm_configs):
            logger.info(f"[{thread_id}] Evaluating study {study_id} with model [{model_idx+1}/{len(self.llm_configs)}]: {llm_config.name}")
            
            # Thread-safe checkpoint checking
            with self.checkpoint_lock:
                checkpoint_key = f"{llm_config.name}:{file_path}"
                if checkpoint_key in self.processed_files:
                    logger.info(f"[{thread_id}] Skipping already processed document: {study_id} (File: {file_name}, Model: {llm_config.name})")
                    continue
            
            # Process document with model
            self.process_document_with_model(file_path, content, llm_config, idx+1, total_files)
            
            # Thread-safe checkpoint saving
            with self.checkpoint_lock:
                self._save_checkpoint(file_path, llm_config.name)

        # Thread-safe results saving
        with self.results_lock:
            logger.info(f"[{thread_id}] Evaluation completed for study {study_id}, saving current results...")
            self.save_results(output_path)
            logger.info(f"[{thread_id}] Evaluation results for study {study_id} saved")

    def process_batch(self, documents: List[str], output_dir: str) -> Dict[str, Any]:
        """
        Process a batch of documents for parallel processing
        
        Args:
            documents: List of document file paths
            output_dir: Output directory for batch results
            
        Returns:
            Dictionary containing batch processing results and statistics
        """
        self.batch_stats['start_time'] = datetime.now()
        self.batch_stats['total_documents'] = len(documents)
        
        batch_id_str = f"batch_{self.batch_id}" if self.batch_id else "single_batch"
        logger.info(f"Starting {batch_id_str} processing with {len(documents)} documents")
        
        os.makedirs(output_dir, exist_ok=True)
        
        batch_results = {
            'batch_id': self.batch_id,
            'processed_documents': [],
            'failed_documents': [],
            'processing_stats': {},
            'errors': []
        }
        
        for idx, file_path in enumerate(documents):
            try:
                result = self.process_single_document(file_path)
                if result['success']:
                    batch_results['processed_documents'].append(result)
                    self.batch_stats['processed_documents'] += 1
                    logger.info(f"Successfully processed document {idx+1}/{len(documents)}: {result['file_name']}")
                else:
                    batch_results['failed_documents'].append(result)
                    self.batch_stats['failed_documents'] += 1
                    logger.error(f"Failed to process document {idx+1}/{len(documents)}: {result['file_name']}")
                    
            except Exception as e:
                error_info = {
                    'file_path': file_path,
                    'error': str(e),
                    'timestamp': datetime.now().isoformat()
                }
                batch_results['errors'].append(error_info)
                self.batch_errors.append(error_info)
                self.batch_stats['failed_documents'] += 1
                logger.error(f"Exception processing document {file_path}: {e}")
        
        # Save batch results
        self.save_batch_results(batch_results, output_dir)
        
        # Generate comprehensive cost report
        if self.cost_analyzer:
            cost_report_paths = self.generate_comprehensive_cost_report(output_dir)
            if cost_report_paths:
                batch_results['cost_report_paths'] = cost_report_paths
                logger.info(f"Cost reports generated: {list(cost_report_paths.keys())}")
        
        # Save error log
        if self.error_handler.error_log:
            error_log_path = os.path.join(output_dir, f"error_log_{self.batch_id}.json")
            self.error_handler.save_error_log(error_log_path)
        
        self.batch_stats['end_time'] = datetime.now()
        batch_results['processing_stats'] = self.batch_stats.copy()
        batch_results['error_summary'] = self.error_handler.get_error_summary()
        
        logger.info(f"Batch processing completed. Processed: {self.batch_stats['processed_documents']}, "
                   f"Failed: {self.batch_stats['failed_documents']}")
        
        if batch_results['error_summary']['total_errors'] > 0:
            logger.warning(f"Total errors encountered: {batch_results['error_summary']['total_errors']}")
            logger.info(f"Error categories: {batch_results['error_summary']['categories']}")
        
        return batch_results

    def process_single_document(self, file_path: str) -> Dict[str, Any]:
        """
        Process a single document with enhanced error handling
        
        Args:
            file_path: Path to the document file
            
        Returns:
            Dictionary containing processing result and metadata
        """
        file_name = os.path.basename(file_path)
        start_time = datetime.now()
        
        result = {
            'file_path': file_path,
            'file_name': file_name,
            'success': False,
            'study_id': None,
            'processing_time': 0,
            'errors': [],
            'model_results': {}
        }
        
        try:
            # Extract text with error handling
            text = self._extract_text_with_retry(file_path)
            if not text:
                result['errors'].append("Failed to extract text from document")
                return result
            
            # Extract study ID with error handling
            study_id = self._extract_study_id_with_retry(text, file_name)
            result['study_id'] = study_id
            
            logger.info(f"Processing document: {file_name} (Study ID: {study_id})")
            logger.info(f"Document length: {len(text)} characters, limit: {self.max_text_length}")
            
            content = {
                'text': text,
                'study_id': study_id,
                'file_name': file_name
            }
            
            # Process with each model
            for model_idx, llm_config in enumerate(self.llm_configs):
                model_result = self._process_document_with_model_safe(
                    file_path, content, llm_config, model_idx + 1, len(self.llm_configs)
                )
                result['model_results'][llm_config.name] = model_result
                
                if not model_result['success']:
                    result['errors'].extend(model_result['errors'])
            
            # Check if at least one model succeeded
            successful_models = [name for name, res in result['model_results'].items() if res['success']]
            result['success'] = len(successful_models) > 0
            
            if result['success']:
                logger.info(f"Document processed successfully with {len(successful_models)} model(s): {file_name}")
            else:
                logger.error(f"All models failed for document: {file_name}")
                
        except Exception as e:
            result['errors'].append(f"Unexpected error: {str(e)}")
            logger.error(f"Unexpected error processing {file_name}: {e}")
        
        finally:
            result['processing_time'] = (datetime.now() - start_time).total_seconds()
        
        return result

    def save_batch_results(self, results: Dict[str, Any], output_dir: str) -> None:
        """
        Save batch processing results to files
        
        Args:
            results: Batch processing results dictionary
            output_dir: Output directory for results
        """
        try:
            # Save main results
            self.save_results(output_dir)
            
            # Save cost analysis if available
            cost_report_path = self.save_cost_report(output_dir)
            if cost_report_path:
                results['cost_report_path'] = cost_report_path
                results['cost_summary'] = self.get_cost_summary()
                results['cost_recommendations'] = self.get_cost_recommendations()
            
            # Save batch metadata
            batch_metadata_file = os.path.join(output_dir, f"batch_metadata_{self.batch_id}.json")
            with open(batch_metadata_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, default=str)
            
            logger.info(f"Batch results saved to {output_dir}")
            
        except Exception as e:
            logger.error(f"Failed to save batch results: {e}")
            raise

    def get_processing_statistics(self) -> Dict[str, Any]:
        """
        Get comprehensive processing statistics including error information
        
        Returns:
            Dictionary containing processing statistics and error summary
        """
        stats = {
            'batch_id': self.batch_id,
            'batch_stats': self.batch_stats.copy(),
            'error_summary': self.error_handler.get_error_summary(),
            'model_results': {}
        }
        
        # Add model-specific statistics
        for model_name, results in self.results.items():
            stats['model_results'][model_name] = {
                'core_assessments': len(results['core']),
                'optional_assessments': len(results['optional']) if self.eval_optional_items else 0
            }
        
        # Calculate processing time if available
        if self.batch_stats['start_time'] and self.batch_stats['end_time']:
            processing_time = (self.batch_stats['end_time'] - self.batch_stats['start_time']).total_seconds()
            stats['total_processing_time_seconds'] = processing_time
            
            if self.batch_stats['processed_documents'] > 0:
                stats['average_time_per_document'] = processing_time / self.batch_stats['processed_documents']
        
        return stats
    
    def get_cost_summary(self) -> Dict[str, Any]:
        """
        Get cost analysis summary
        
        Returns:
            Dictionary containing cost summary or empty dict if cost analyzer not available
        """
        if self.cost_analyzer:
            return self.cost_analyzer.get_cost_summary()
        else:
            return {
                'cost_tracking_enabled': False,
                'message': 'Cost tracking not available'
            }
    
    def save_cost_report(self, output_dir: str) -> Optional[str]:
        """
        Save detailed cost report to output directory
        
        Args:
            output_dir: Directory to save the cost report
            
        Returns:
            Path to saved cost report file, or None if cost analyzer not available
        """
        if not self.cost_analyzer:
            logger.warning("Cost analyzer not available, cannot save cost report")
            return None
        
        try:
            cost_log_path = os.path.join(output_dir, f"cost_analysis_{self.cost_analyzer.session_id}.json")
            self.cost_analyzer.save_detailed_log(cost_log_path)
            logger.info(f"Cost analysis report saved to: {cost_log_path}")
            return cost_log_path
        except Exception as e:
            logger.error(f"Failed to save cost report: {e}")
            return None
    
    def get_cost_recommendations(self) -> List[str]:
        """
        Get cost optimization recommendations
        
        Returns:
            List of recommendation strings
        """
        if self.cost_analyzer:
            return self.cost_analyzer.generate_recommendations()
        else:
            return ["Cost tracking not available - enable cost analyzer for recommendations"]
    
    def generate_comprehensive_cost_report(self, output_dir: str, 
                                         currencies: Optional[List[str]] = None) -> Dict[str, str]:
        """
        Generate comprehensive cost report in multiple formats
        
        Args:
            output_dir: Directory to save the cost reports
            currencies: List of currencies to include (default: ['USD', 'EUR', 'CNY'])
            
        Returns:
            Dictionary mapping format names to file paths, or empty dict if cost analyzer not available
        """
        if not self.cost_analyzer:
            logger.warning("Cost analyzer not available, cannot generate comprehensive cost report")
            return {}
        
        try:
            from cost_reporter import CostReporter
            cost_reporter = CostReporter(self.cost_analyzer)
            
            # Generate comprehensive report
            currencies_list = currencies or ['USD', 'EUR', 'CNY']
            report_paths = cost_reporter.generate_comprehensive_report(output_dir, currencies_list)
            
            logger.info(f"Comprehensive cost report generated with {len(report_paths)} formats")
            for format_name, path in report_paths.items():
                logger.info(f"  {format_name.upper()}: {path}")
            
            return report_paths
            
        except ImportError as e:
            logger.error(f"Failed to import CostReporter: {e}")
            return {}
        except Exception as e:
            logger.error(f"Failed to generate comprehensive cost report: {e}")
            return {}
    
    def display_cost_summary_in_progress(self) -> str:
        """
        Get formatted cost summary for progress display
        
        Returns:
            Formatted string with current cost information
        """
        if not self.cost_analyzer:
            return "💰 Cost tracking: Not available"
        
        try:
            cost_summary = self.cost_analyzer.get_cost_summary()
            
            lines = [
                f"💰 Total Cost: ${cost_summary['total_cost_usd']:.4f} USD",
                f"🔢 Total Tokens: {cost_summary['total_tokens']:,}",
                f"📞 API Calls: {cost_summary['total_api_calls']:,}"
            ]
            
            # Add model breakdown
            if cost_summary['model_summaries']:
                lines.append("📊 Model Breakdown:")
                for model_summary in cost_summary['model_summaries']:
                    model_name = model_summary['model']
                    cost = model_summary['total_cost_usd']
                    tokens = model_summary['total_tokens']
                    lines.append(f"  • {model_name}: ${cost:.4f} ({tokens:,} tokens)")
            
            # Add top recommendation
            recommendations = self.cost_analyzer.generate_recommendations()
            if recommendations:
                lines.append(f"💡 Tip: {recommendations[0]}")
            
            return "\n".join(lines)
            
        except Exception as e:
            logger.error(f"Failed to generate cost summary display: {e}")
            return "💰 Cost tracking: Error retrieving data"

    def should_continue_processing(self) -> bool:
        """
        Determine if processing should continue based on error patterns
        
        Returns:
            True if processing should continue, False if it should stop
        """
        error_summary = self.error_handler.get_error_summary()
        
        # Stop if too many critical errors
        critical_errors = error_summary['severities'].get('critical', 0)
        if critical_errors > 0:
            logger.critical("Critical errors detected, stopping processing")
            return False
        
        # Stop if too many high severity errors
        high_errors = error_summary['severities'].get('high', 0)
        total_errors = error_summary['total_errors']
        
        if total_errors > 0:
            high_error_ratio = high_errors / total_errors
            if high_error_ratio > 0.5 and total_errors > 5:  # More than 50% high severity errors
                logger.error("Too many high severity errors, stopping processing")
                return False
        
        return True

    def extract_study_id(self, document_text: str, filename: str, llm_config: LLMConfig) -> str:

        prompt = f"""
        Your task is to extract the study identifier in the format "Surname, YYYY" where Surname is the first author's surname and YYYY is the publication year (4 digits).

        First, try to extract this information from the document text. Look for:
        1. The surname of the first author
        2. The publication year (4 digits)

        If you cannot find this information in the document text, try to extract it from the filename: "{filename}"

        Return ONLY the formatted string as Surname, YYYY.

        Document text (excerpt):
        {document_text[:4000]}

        Example response formats:
        Smith, 2020
        Johnson, 2019

        Return the content in pure ID form without additional symbols such as quotation marks.
        """
        if self.cost_analyzer:
            response, usage_info = llm_config.generate_response_with_usage(prompt)
            response = response.strip()
            if usage_info:
                self.cost_analyzer.track_usage(
                    model=llm_config.model_name,
                    input_tokens=usage_info['input_tokens'],
                    output_tokens=usage_info['output_tokens'],
                    document_name=filename,
                    operation="study_id_extraction"
                )
        else:
            response = llm_config.generate_response(prompt).strip()
        
        logger.debug(f"Extracted Study ID response: {response}")

        def is_valid_study_id(response):
            # 先清理响应，移除可能的引号
            cleaned_response = response.strip().strip('"').strip("'")
            # 允许姓氏包含字母、空格、撇号和连字符，后跟逗号、可选空格和4位数字
            pattern = r'^[A-Z][a-zA-Z\s\'-]+,\s*\d{4}$'
            return bool(re.match(pattern, cleaned_response))

        # 清理响应
        cleaned_response = response.strip().strip('"').strip("'")
        
        if is_valid_study_id(response):
            return cleaned_response
        else:
            logger.warning(f"Study ID format mismatch: '{response}', cleaned: '{cleaned_response}'")
            # 尝试从文件名中提取作为备选方案
            filename_match = re.search(r'([A-Z][a-zA-Z\s\'-]+).*?(\d{4})', filename)
            if filename_match:
                surname = filename_match.group(1).strip()
                year = filename_match.group(2)
                fallback_id = f"{surname}, {year}"
                logger.info(f"Extracted fallback study ID from filename: {fallback_id}")
                return fallback_id
            return "Unknown, 0000"

    def process_document_with_model(self, file_path: str, content: dict, llm_config: LLMConfig, current_idx: int, total_docs: int):
        """Process a single document using a specific model"""
        file_name = content['file_name']
        text = content['text']
        study_id = content['study_id']
        
        # Extract filename without extension
        base_file_name = os.path.splitext(file_name)[0]
        text_snippet = text[:self.max_text_length]

        # 1. Process core items
        logger.info(f"Evaluating core items for study {study_id} using model {llm_config.name} [{current_idx}/{total_docs}]")
        core_prompt = self.generate_core_prompt(text_snippet, self.output_mode)
        start_time = time.time()
        
        if self.output_mode == 'json':
            if self.cost_analyzer:
                core_response, usage_info = llm_config.generate_structured_response_with_usage(
                    core_prompt, 
                    response_format={"type": "json_schema", "json_schema": {"name": "core_assessment", "schema": CoreAssessmentResponse.model_json_schema()}}
                )
                if usage_info:
                    self.cost_analyzer.track_usage(
                        model=llm_config.model_name,
                        input_tokens=usage_info['input_tokens'],
                        output_tokens=usage_info['output_tokens'],
                        document_name=file_name,
                        operation="core_assessment"
                    )
            else:
                core_response = llm_config.generate_structured_response(
                    core_prompt, 
                    response_format={"type": "json_schema", "json_schema": {"name": "core_assessment", "schema": CoreAssessmentResponse.model_json_schema()}}
                )
            core_results = self.parse_core_structured_response(study_id, core_response, base_file_name)
        
        elif self.output_mode == 'table':
            if self.cost_analyzer:
                core_response, usage_info = llm_config.generate_response_with_usage(core_prompt)
                if usage_info:
                    self.cost_analyzer.track_usage(
                        model=llm_config.model_name,
                        input_tokens=usage_info['input_tokens'],
                        output_tokens=usage_info['output_tokens'],
                        document_name=file_name,
                        operation="core_assessment"
                    )
            else:
                core_response = llm_config.generate_response(core_prompt)
            core_results = self.parse_core_table_response(study_id, core_response, base_file_name)
        else:
            logger.error(f"Unsupported output mode: {self.output_mode}")
            return
        core_elapsed_time = time.time() - start_time
        logger.info(f"Core items evaluation took: {core_elapsed_time:.2f} seconds")
        if core_results:
            logger.info(f"Successfully parsed core items evaluation results for {study_id}")
            self.results[llm_config.name]['core'].extend(core_results)
        else:
            logger.error(f"Failed to parse core items evaluation results for {study_id}")
        
        # 2. Process optional items (if enabled)
        if self.eval_optional_items:
            logger.info(f"Evaluating optional items for study {study_id} using model {llm_config.name} [{current_idx}/{total_docs}]")
            optional_prompt = self.generate_optional_prompt(text_snippet, self.output_mode)
            start_time = time.time()
            if self.output_mode == 'json':
                if self.cost_analyzer:
                    optional_response, usage_info = llm_config.generate_structured_response_with_usage(
                        optional_prompt,
                        response_format={"type": "json_schema", "json_schema": {"name": "optional_assessment", "schema": OptionalAssessmentResponse.model_json_schema()}}
                    )
                    if usage_info:
                        self.cost_analyzer.track_usage(
                            model=llm_config.model_name,
                            input_tokens=usage_info['input_tokens'],
                            output_tokens=usage_info['output_tokens'],
                            document_name=file_name,
                            operation="optional_assessment"
                        )
                else:
                    optional_response = llm_config.generate_structured_response(
                        optional_prompt,
                        response_format={"type": "json_schema", "json_schema": {"name": "optional_assessment", "schema": OptionalAssessmentResponse.model_json_schema()}}
                    )
                optional_results = self.parse_optional_structured_response(study_id, optional_response, base_file_name)
            elif self.output_mode == 'table':
                if self.cost_analyzer:
                    optional_response, usage_info = llm_config.generate_response_with_usage(optional_prompt)
                    if usage_info:
                        self.cost_analyzer.track_usage(
                            model=llm_config.model_name,
                            input_tokens=usage_info['input_tokens'],
                            output_tokens=usage_info['output_tokens'],
                            document_name=file_name,
                            operation="optional_assessment"
                        )
                else:
                    optional_response = llm_config.generate_response(optional_prompt)
                optional_results = self.parse_optional_table_response(study_id, optional_response, base_file_name)
            else:
                logger.error(f"Unsupported output mode: {self.output_mode}")
                return
            optional_elapsed_time = time.time() - start_time
            logger.info(f"Optional items evaluation took: {optional_elapsed_time:.2f} seconds")
            if optional_results:
                logger.info(f"Successfully parsed optional items evaluation results for {study_id}")
                self.results[llm_config.name]['optional'].extend(optional_results)
            else:
                logger.error(f"Failed to parse optional items evaluation results for {study_id}")


    def generate_core_prompt(self, text: str, output_mode: str) -> str:
        json_instructions = """
## RESPONSE FORMAT INSTRUCTIONS (JSON)
You must return your assessment as a structured JSON response following this exact format:
{{
  "study_id": "[Author surname, Year]",
  "assessments": [
    {{
      "item_number": 1,
      "step1_decision": "[Step 1 decision]",
      "step2_decision": "[Step 2 decision]", 
      "reason": "[Your reasoning for this judgment]",
      "quote": "[Direct quote from text supporting judgment]"
    }},
    ... (continue for all 6 items)
  ]
}}
For Item 6, use "step1_decision" for the percentage value and "step2_decision" will be automatically determined.
"""
        table_instructions = """
## RESPONSE FORMAT INSTRUCTIONS (TABLE)
You MUST return your assessment in a markdown table format. The table must have these exact 5 columns: `Item Number`, `Step 1 Decision`, `Step 2 Decision`, `Reason`, `Quote`.
- Each row corresponds to one assessment item (1 through 6).
- Start the response DIRECTLY with the first row of the table. Do NOT include the markdown table header (e.g., `| Item Number | ... |`).
- Use the pipe `|` character as a separator.
- For Item 6, put the extracted percentage or "Not reported" in the `Step 1 Decision` column and leave the `Step 2 Decision` column empty. It will be calculated later.
- Ensure the `Quote` is on a single line; replace newlines within the quote with a space.
- Do not add any text before or after the table.
Example of a valid response:
| 1 | Probably no | Probably high | The study was described as randomized but provided no details on the method of sequence generation. | The trial was described as a randomized, double-blind, placebo-controlled study. |
| 2 | Probably yes | Probably low | The study used sequentially numbered opaque sealed envelopes. | Allocation was concealed using sequentially numbered, opaque, sealed envelopes. |
| 3 | Definitely yes | Definitely low | The report explicitly stated that participants were blinded. | All participants were blinded to the treatment allocation. |
| 4 | Definitely yes | Definitely low | The report explicitly stated that healthcare providers were blinded. | The study was double-blind, so providers were unaware of the treatment. |
| 5 | Probably yes | Probably low | The study was described as double-blind, implying outcome assessors were blinded. | The double-blind nature of the trial ensured that outcome assessors were not aware of group assignment. |
| 6 | 5.2% | | The analysis excluded 10 out of 192 randomized participants due to protocol violations. | A total of 192 participants were randomized, but 10 were excluded from the primary analysis. |
"""
        
        format_instructions = table_instructions if output_mode == 'table' else json_instructions
        
        base_prompt = f"""
# ROBUST-RCT Risk of Bias Assessment

## Introduction and Role
You are a methodological expert applying the Risk Of Bias instrument for Use in SysTematic reviews for Randomized Controlled Trials (ROBUST-RCT). Your task is to systematically assess potential sources of bias in randomized trials using the framework and criteria specified below, basing your assessment solely on information presented in the trial report.

## Assessment Structure
For each of the six core items (Items 1-5), follow this two-step process:
- Step 1: Evaluate what happened (whether methodological safeguards were implemented)
- Step 2: Judge the risk of bias based on what happened in Step 1

For Items 1-5, use these response options:
- Step 1: "Definitely yes," "Probably yes," "Probably no," or "Definitely no"
- Step 2: "Definitely low," "Probably low," "Probably high," or "Definitely high" risk of bias

For Item 6, your task is different:
- Step 1: Extract the percentage of randomized participants who were NOT INCLUDED IN THE ANALYSIS. Return ONLY the percentage value (e.g., "5.2%") or "Not reported" if the study doesn't report this information.
  - Focus specifically on participants who were completely excluded from the analysis, not those who dropped out but still had data included (e.g., in intention-to-treat analysis)
  - Look for information about participants excluded from analysis due to: protocol violations, loss to follow-up with no data available, withdrawal with data not analyzed, or other reasons for complete exclusion from results
  - Do not calculate this percentage yourself or make assumptions if not clearly reported
  - Do not provide the raw numbers, only the percentage
  - If multiple analyses are reported, focus on the primary outcome analysis
  - If completely unclear, just write "Not reported"

## Core Assessment Items

### Item 1: Random Sequence Generation
- Step 1: Was the allocation sequence adequately generated?
  - Definitely yes: Trial explicitly stated use of an adequate method (random number table, random number generator, throwing dice, drawing of lots, minimization)
  - Probably yes: Trial described as "randomized" and either: 1) mentioned simple/block/stratified randomization, or 2) described allocation concealment method
  - Probably no: Trial described as "randomized" without further details that don't meet "Probably yes" criteria
  - Definitely no: Trial used non-randomized allocation (quasi-randomization) such as allocation by date of birth/admission, record number, alternation, clinician/participant decision, or test results
- Step 2: Judge risk of bias related to sequence generation (follows same criteria as Step 1)
**IMPORTANT**: If the study does not explicitly provide clear details about the methodological aspect being assessed, you MUST use "Probably no" for Step 1. Do not make assumptions based on study type or context. Only explicit statements in the text should lead to "yes" judgments.

### Item 2: Allocation Concealment
- Step 1: Was allocation adequately concealed?
  - Definitely yes: Trial used central allocation, pharmacy-controlled randomization, or sequentially numbered opaque sealed envelopes with evidence of sequential opening
  - Probably yes: Trial used sequentially numbered opaque sealed envelopes without further details, OR drug trial with blinded participants/providers, OR explicitly stated concealment without details, OR used minimization approach
  - Probably no: Unblinded drug trial or non-drug trial without allocation concealment details, OR used envelopes/containers with unclear safeguards
  - Definitely no: Trial used an open allocation schedule OR used quasi-randomization methods
- Step 2: Judge risk of bias related to allocation concealment (follows same criteria as Step 1)
**IMPORTANT**: If the study does not explicitly provide clear details about the methodological aspect being assessed, you MUST use "Probably no" for Step 1. Do not make assumptions based on study type or context. Only explicit statements in the text should lead to "yes" judgments.

### Item 3: Blinding of Participants
- Step 1: Were participants blinded?
  - Definitely yes: Trial explicitly stated participants were blinded
  - Probably yes: No explicit statement but: placebo-controlled drug trial, OR active-control drug trial mentioning "double dummy"/identical medications, OR described as blinded with likely participant blinding
  - Probably no: No explicit statement and: active-control drug trial without "double dummy"/identical medications, OR non-drug trial, OR single-blinded with participants likely not blinded
  - Definitely no: Trial explicitly stated participants were not blinded, OR described as "open-label"/"unblinded"
- Step 2: Judge risk of bias related to blinding of participants
  - Consider: 1) Were participants blinded? 2) If unblinded, how likely did this influence outcomes through expectations or co-interventions?
  - Definitely low: Participants definitely blinded, OR unblinding very unlikely to influence outcome
  - Probably low: Participants probably blinded, OR unblinding unlikely to influence outcome
  - Probably high: Participants not blinded AND unblinding likely influenced outcome through expectations or co-interventions
  - Definitely high: Participants not blinded AND unblinding very likely influenced outcome through expectations or co-interventions

### Item 4: Blinding of Healthcare Providers
- Step 1: Were healthcare providers blinded?
  - Follow same criteria as Item 3, Step 1, applied to healthcare providers
- Step 2: Judge risk of bias related to blinding of healthcare providers
  - Consider: 1) Were providers blinded? 2) If unblinded, how likely did provider-initiated co-interventions influence outcomes?
  - Definitely low: Providers definitely blinded
  - Probably low: Providers probably blinded, OR provider-initiated co-interventions unlikely to influence outcome
  - Probably high: Providers not blinded AND co-interventions likely influenced outcome
  - Definitely high: Providers not blinded AND documented dissimilarity in co-interventions between groups

### Item 5: Blinding of Outcome Assessors
- Step 1: Were outcome assessors blinded?
  - Follow same criteria as Item 3, Step 1, applied to outcome assessors (note who the assessors are for each outcome)
- Step 2: Judge risk of bias related to blinding of outcome assessors
  - Consider: 1) Were assessors blinded? 2) If unblinded, how subjective was the outcome assessment?
  - Definitely low: Assessors definitely blinded, OR outcome is all-cause mortality
  - Probably low: Assessors probably blinded, OR outcome assessment involves minimal judgment
  - Probably high: Assessors not blinded AND outcome assessment involves some judgment
  - Definitely high: Assessors not blinded AND outcome assessment involves considerable judgment

### Item 6: Outcome Data Not Included in Analysis
- Step 1: What percentage of randomized participants were NOT INCLUDED in the analysis?
  - Return ONLY the percentage (e.g., "5.2%") or "Not reported" if unclear
  - Focus on participants completely excluded from analysis, not those who dropped out but had data included in intention-to-treat analysis
  - Do not include raw numbers or calculations
  - If multiple outcomes have different exclusion rates, report the highest percentage

{format_instructions}

For Item 6, use "step1_decision" for the percentage value and "step2_decision" will be automatically determined.
Study Content:

Study Content:
{text}
"""
        return base_prompt


    def generate_optional_prompt(self, text: str, output_mode: str) -> str:
        json_instructions = """
## RESPONSE FORMAT INSTRUCTIONS (JSON)
You must return your assessment as a structured JSON response following this exact format:
{{
  "study_id": "[Author surname, Year]",
  "assessments": [
    {{
      "item_number": 1,
      "decision": "[Your risk of bias judgment]",
      "reason": "[Your reasoning for this judgment]",
      "quote": "[Direct quote from text supporting judgment]"
    }},
    ... (continue for all 8 optional items)
  ]
}}
"""
        table_instructions = """
## RESPONSE FORMAT INSTRUCTIONS (TABLE)
You MUST return your assessment in a markdown table format. The table must have these exact 4 columns: `Item Number`, `Decision`, `Reason`, `Quote`.
- Each row corresponds to one optional assessment item (1 through 8).
- Start the response DIRECTLY with the first row of the table. Do NOT include the markdown table header.
- Use the pipe `|` character as a separator.
- Ensure the `Quote` is on a single line; replace newlines within the quote with a space.
- Do not add any text before or after the table.
Example of a valid response:
| 1 | Probably low | Baseline characteristics such as age and gender were well-balanced between groups as shown in Table 1. | Table 1 shows similar baseline demographics for both the intervention and control groups. |
| 2 | Probably low | The trial was double-blinded, making differential co-interventions unlikely. | The double-blind design was maintained throughout the study. |
"""
        
        format_instructions = table_instructions if output_mode == 'table' else json_instructions
        base_prompt = f"""
# ROBUST-RCT Optional Items Risk of Bias Assessment

## Introduction
You are a methodological expert continuing your assessment of a randomized controlled trial. Having completed the core items evaluation, you will now assess optional items that may be relevant for this study.

## Assessment Instructions
For each optional item, follow a step-by-step approach:
1. Carefully read the study to identify relevant information
2. Consider the specific criteria for each optional item
3. Make a judgment based solely on the information in the study
4. Document your reasoning and provide direct quotes from the study

## Important Guidelines for Making Judgments:
- When direct information is not available, use indirect information from the study to inform your judgment
- All indirect information must come from the original text; do not make unsupported assumptions
- When making judgments based on indirect information or when information is missing, use "Probably high" or "Probably low" risk (never "Definitely")
- When no information is available (neither direct nor indirect), explain this and default to "Probably high" risk
- Never use "Not applicable" as a final judgment when information is missing; instead explain why the item is relevant but cannot be assessed due to missing information

## Optional Items to Assess
### Item 1: Balance of baseline prognostic factors between groups
Think through: Does the study report baseline characteristics? Are important prognostic factors balanced between intervention groups? If imbalances exist, are they substantial enough to potentially bias results?
- Definitely low risk: Clear evidence of balance in all important prognostic factors
- Probably low risk: Evidence suggests balance in most important factors, or any imbalance unlikely to bias results
- Probably high risk: Evidence suggests imbalance in some important factors that may bias results, or insufficient information to assess balance
- Definitely high risk: Clear evidence of substantial imbalance in important prognostic factors

### Item 2: Balance of co-interventions between groups in blinded trials
Think through: In this blinded trial, were important co-interventions that could influence outcomes balanced between groups? Were any documented differences in co-interventions unrelated to the intervention effects?

### Item 3: Different outcome assessment/data collection between groups
Think through: Were the methods for measuring outcomes or collecting data identical across intervention groups? Were there systematic differences in assessment approaches?

### Item 4: Different follow-up time/frequency/intensity between groups
Think through: Did the study apply the same follow-up schedule and assessment frequency to all intervention groups? Were there systematic differences in follow-up intensity?

### Item 5: Validity of outcome measurement methods
Think through: Were the methods used to measure outcomes valid and reliable? Were established, validated measurement instruments used appropriately?

### Item 6: As-treated analysis concerns
Think through: Did the study conduct an as-treated analysis (analyzing participants according to intervention received rather than allocated)? What percentage of participants were analyzed in groups other than those to which they were randomized?
- If the study explicitly mentions no as-treated analysis was conducted, state this and assess as "Probably low" risk
- If the study doesn't mention as-treated analysis but describes analysis by original allocation, use this indirect information to assess as "Probably low" risk
- If no information is provided about analysis approach, explain this and assess as "Probably high" risk
- Never state "Not applicable" just because information is missing

### Item 7: Selective outcome reporting
Think through: Does the study report all pre-specified outcomes as described in the protocol or methods? Is there evidence that results were selectively reported based on their magnitude or significance?

### Item 8: Early trial termination for benefit
Think through: Was the trial stopped early due to benefits observed in interim analyses? If so, what impact might this have on the reliability of the results?

{format_instructions}

Study Content:
{text}
"""
        return base_prompt

    def parse_core_table_response(self, study_id: str, response: str, file_name: str):
        """Parses the table response for core item assessment"""
        results = []
        # Clean the response by removing possible code block markers
        cleaned_response = re.sub(r'```(markdown)?', '', response, flags=re.IGNORECASE).strip()
        lines = cleaned_response.strip().split('\n')
        for i, line in enumerate(lines):
            line = line.strip()
            if not line.startswith('|') or not line.endswith('|'):
                continue
            parts = [p.strip() for p in line.strip('|').split('|')]
            if len(parts) != 5:
                logger.warning(f"Failed to parse table row (not 5 fields): {line} in {study_id}")
                continue
            try:
                item_num_str, step1_original, step2_original, reason_original, quote = parts
                item_num = int(item_num_str)
                # Domain mapping (consistent with JSON parsing function)
                domain_name = {
                    1: "1. Random sequence generation", 2: "2. Allocation concealment",
                    3: "3. Blinding of participants", 4: "4. Blinding of healthcare providers",
                    5: "5. Blinding of outcome assessors", 6: "6. Outcome data not included in analysis"
                }.get(item_num, f"Unknown Core Item {item_num}")
                # Correction logic (consistent with JSON parsing function)
                if item_num <= 2 and self.detect_insufficient_evidence_patterns(reason_original):
                    step1_corrected = "Probably no"
                    step2_corrected = "Probably high"
                    reason_corrected = f"[CORRECTED: Evidence insufficient] {reason_original} [Original: {step1_original}/{step2_original}]"
                else:
                    step1_corrected = step1_original
                    step2_corrected = step2_original
                    reason_corrected = reason_original
                result_item = {
                    "File name": file_name, "Study": study_id, "Domain": domain_name,
                    "Step 1": step1_corrected, "Step 2": step2_corrected,
                    "Reason": reason_corrected, "Quote": quote
                }
                # Special handling for Item 6
                if item_num == 6:
                    step1_value = step1_original
                    result_item["Step 1"] = step1_value
                    result_item["Step 2"] = self.determine_domain6_risk(step1_value)
                results.append(result_item)

            except (ValueError, IndexError) as e:
                logger.error(f"Error parsing table row: {e} - Line: {line} in {study_id}")
                continue

        if not results:
            logger.error(f"No valid rows parsed from table response: {study_id}")
            error_file = f"error_table_core_{study_id.replace(', ', '_')}.txt"
            with open(error_file, 'w', encoding='utf-8') as f:
                f.write(response)
            logger.error(f"Full core item table response saved to {error_file}")
        return results

    def parse_optional_table_response(self, study_id: str, response: str, file_name: str):
        """Parses the table response for optional items assessment"""
        results = []
        cleaned_response = re.sub(r'```(markdown)?', '', response, flags=re.IGNORECASE).strip()

        lines = cleaned_response.strip().split('\n')
        for i, line in enumerate(lines):
            line = line.strip()
            if not line.startswith('|') or not line.endswith('|'):
                continue
            parts = [p.strip() for p in line.strip('|').split('|')]
            if len(parts) != 4:
                logger.warning(f"Failed to parse optional item table row (not 4 columns): {line} in {study_id}")
                continue

            try:
                item_num_str, decision, reason, quote = parts
                item_num = int(item_num_str)
                # Domain name mapping (consistent with JSON parsing function)
                domain_name = {
                    1: "7. Balance of baseline prognostic factors",
                    2: "8. Balance of co-interventions in blinded trials",
                    3: "9. Different outcome assessment between groups",
                    4: "10. Different follow-up between groups",
                    5: "11. Validity of outcome measurement methods",
                    6: "12. As-treated analysis concerns",
                    7: "13. Selective outcome reporting",
                    8: "14. Early trial termination for benefit"
                }.get(item_num, f"Unknown Optional Item {item_num}")
                result_item = {
                    "File name": file_name,
                    "Study": study_id,
                    "Domain": domain_name,
                    "Step 1": "Not available",
                    "Step 2": decision,
                    "Reason": reason,
                    "Quote": quote
                }
                results.append(result_item)
            except (ValueError, IndexError) as e:
                logger.error(f"Error parsing optional table row: {e} - Line: {line} in {study_id}")
                continue

        if not results:
            logger.error(f"No valid optional item rows parsed from table response: {study_id}")
            error_file = f"error_table_optional_{study_id.replace(', ', '_')}.txt"
            with open(error_file, 'w', encoding='utf-8') as f:
                f.write(response)
            logger.error(f"Full optional items table response saved to {error_file}")
        return results

    def determine_domain6_risk(self, percentage_str: str) -> str:
        """Determine the risk level for domain 6 based on the percentage."""
        if percentage_str == "Not reported":
            return self.domain6_default

        try:
            # Try to extract the numerical value from the percentage string
            percentage = float(percentage_str.replace('%', '').strip())

            # Determine the risk level based on thresholds
            if percentage < self.domain6_thresholds["definitely_low"]:
                return "Definitely low"
            elif percentage < self.domain6_thresholds["probably_low"]:
                return "Probably low"
            elif percentage < self.domain6_thresholds["probably_high"]:
                return "Probably high"
            else:
                return "Definitely high"
        except (ValueError, TypeError):
            logger.warning(f"Could not parse percentage string: '{percentage_str}', using default risk level: {self.domain6_default}")
            return self.domain6_default

    def parse_core_structured_response(self, study_id: str, response: str, file_name: str):
        """Parse structured response from core item evaluation"""
        try:
            import json
            data = json.loads(response)
            
            results = []
            
            for assessment in data.get('assessments', []):
                item_num = assessment.get('item_number')
                
                # 域名映射
                domain_name = ""
                if item_num == 1:
                    domain_name = "1. Random sequence generation"
                elif item_num == 2:
                    domain_name = "2. Allocation concealment"
                elif item_num == 3:
                    domain_name = "3. Blinding of participants"
                elif item_num == 4:
                    domain_name = "4. Blinding of healthcare providers"
                elif item_num == 5:
                    domain_name = "5. Blinding of outcome assessors"
                elif item_num == 6:
                    domain_name = "6. Outcome data not included in analysis"
                
                # 获取原始评估结果
                step1_original = assessment.get('step1_decision', '')
                step2_original = assessment.get('step2_decision', '')
                reason_original = assessment.get('reason', '')
                # 检查是否需要修正（仅对条目1-x进行修正）
                if item_num <= 2 and self.detect_insufficient_evidence_patterns(reason_original):
                    step1_corrected = "Probably no"
                    step2_corrected = "Probably high"
                    reason_corrected = f"[CORRECTED: Evidence insufficient] {reason_original} [Original: {step1_original}/{step2_original}]"
                else:
                    step1_corrected = step1_original
                    step2_corrected = step2_original
                    reason_corrected = reason_original
                result_item = {
                    "File name": file_name,
                    "Study": study_id,
                    "Domain": domain_name,
                    "Step 1": step1_corrected,
                    "Step 2": step2_corrected,
                    "Reason": reason_corrected,
                    "Quote": assessment.get('quote', '')
                }
                
                # 特殊处理 Item 6
                if item_num == 6:
                    step1_value = assessment.get('step1_decision', '')
                    result_item["Step 1"] = step1_value
                    result_item["Step 2"] = self.determine_domain6_risk(step1_value)
                
                results.append(result_item)
            
            return results
            
        except Exception as e:
            logger.error(f"Failed to parse structured core items response: {e}")
            # Save the full response content to a file for debugging
            error_file = f"error_structured_core_{study_id.replace(', ', '_')}.txt"
            with open(error_file, 'w', encoding='utf-8') as f:
                f.write(response)
            logger.error(f"Full structured core items response has been saved to {error_file}")
            return []

    def parse_optional_structured_response(self, study_id: str, response: str, file_name: str):
        """Parse structured response from optional item evaluation"""
        try:
            import json
            data = json.loads(response)
            
            results = []
            
            for assessment in data.get('assessments', []):
                item_num = assessment.get('item_number')
                
                # 可选条目域名映射
                domain_name = ""
                if item_num == 1:
                    domain_name = "7. Balance of baseline prognostic factors"
                elif item_num == 2:
                    domain_name = "8. Balance of co-interventions in blinded trials"
                elif item_num == 3:
                    domain_name = "9. Different outcome assessment between groups"
                elif item_num == 4:
                    domain_name = "10. Different follow-up between groups"
                elif item_num == 5:
                    domain_name = "11. Validity of outcome measurement methods"
                elif item_num == 6:
                    domain_name = "12. As-treated analysis concerns"
                elif item_num == 7:
                    domain_name = "13. Selective outcome reporting"
                elif item_num == 8:
                    domain_name = "14. Early trial termination for benefit"
                
                result_item = {
                    "File name": file_name,
                    "Study": study_id,
                    "Domain": domain_name,
                    "Step 1": "Not available",
                    "Step 2": assessment.get('decision', ''),
                    "Reason": assessment.get('reason', ''),
                    "Quote": assessment.get('quote', '')
                }
                
                results.append(result_item)
            
            return results
            
        except Exception as e:
            logger.error(f"Failed to parse structured optional entries response: {e}")
            # Save the full response to a file for debugging
            error_file = f"error_structured_optional_{study_id.replace(', ', '_')}.txt"
            with open(error_file, 'w', encoding='utf-8') as f:
                f.write(response)
            logger.error(f"Full structured optional entries response saved to {error_file}")
            return []

    def _clean_text_for_excel(self, text):
        """Clean characters not supported by Excel from text"""
        if not isinstance(text, str):
            return str(text) if text is not None else ""
        
        # 移除ASCII控制字符（保留制表符、换行符、回车符）
        text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
        
        # 移除其他可能有问题的Unicode字符
        text = ''.join(char for char in text if ord(char) >= 32 or char in '\t\n\r')
        
        # 替换可能有问题的特殊字符
        text = text.replace('\x00', '')  # 空字符
        
        # 限制文本长度（Excel单元格有字符限制）
        if len(text) > 32767:  # Excel单元格最大字符数
            text = text[:32764] + "..."
        
        return text.strip()


    def save_results(self, output_path: str):
        """Save the results to Excel, integrate all model results into one sheet, and highlight Step 2 differences"""
        os.makedirs(output_path, exist_ok=True)
        output_file = os.path.join(output_path, "rob_results.xlsx")
        logger.info(f"Creating/updating Excel file: {output_file}")

        # Check if there are any results
        has_any_results = False
        for model_name, model_results in self.results.items():
            if model_results['core'] or (self.eval_optional_items and model_results['optional']):
                has_any_results = True
                break
        
        if not has_any_results:
            logger.warning("No evaluation results to save!")
            return

        # 获取所有研究ID和域
        all_studies = set()
        all_domains = set()
        for model_results in self.results.values():
            for category in ['core', 'optional']:
                for result in model_results[category]:
                    all_studies.add(result["Study"])
                    all_domains.add(result["Domain"])
        
        logger.info(f"Currently there are {len(all_studies)} studies that have been evaluated")

        # 预定义domain顺序
        core_domains = [
            "1. Random sequence generation",
            "2. Allocation concealment", 
            "3. Blinding of participants",
            "4. Blinding of healthcare providers",
            "5. Blinding of outcome assessors",
            "6. Outcome data not included in analysis"
        ]
        
        optional_domains = [
            "7. Balance of baseline prognostic factors",
            "8. Balance of co-interventions in blinded trials",
            "9. Different outcome assessment between groups",
            "10. Different follow-up between groups",
            "11. Validity of outcome measurement methods",
            "12. As-treated analysis concerns",
            "13. Selective outcome reporting",
            "14. Early trial termination for benefit"
        ] if self.eval_optional_items else []
        
        domains = core_domains + optional_domains

        # 收集每个研究对应的文件名
        file_study_pairs = set()
        for model_name in self.results.keys():
            for category in ['core', 'optional']:
                for result in self.results[model_name][category]:
                    file_study_pairs.add((result.get("File name", "Unknown"), result["Study"]))

        # 获取模型名称列表
        model_names = list(self.results.keys())
        
        # 构建综合数据
        integrated_data = []
        
        for file_name, study in sorted(file_study_pairs):
            for domain in domains:
                        
                # 初始化行数据
                row_data = {
                    "File name": self._clean_text_for_excel(file_name),
                    "Study": self._clean_text_for_excel(study),
                    "Domain": self._clean_text_for_excel(domain)
                }
                
                # 判断是否为核心条目
                is_core_item = domain in core_domains
                
                # 为每个模型添加数据
                for model_name in model_names:
                    category = 'core' if is_core_item else 'optional'
                    
                    # 查找该模型对该研究该域的评估结果
                    model_result = None
                    for result in self.results[model_name][category]:
                        if result["Study"] == study and result["Domain"] == domain:
                            model_result = result
                            break
                    
                    if model_result:
                        if is_core_item:
                            # 核心条目有Step 1和Step 2
                            row_data[f"{model_name} (Step 1)"] = self._clean_text_for_excel(model_result.get("Step 1", ""))
                            row_data[f"{model_name} (Step 2)"] = self._clean_text_for_excel(model_result.get("Step 2", ""))
                        else:
                            # 可选条目没有Step 1
                            row_data[f"{model_name} (Step 1)"] = "NA"
                            row_data[f"{model_name} (Step 2)"] = self._clean_text_for_excel(model_result.get("Step 2", ""))
                        
                        row_data[f"{model_name} (Reason)"] = self._clean_text_for_excel(model_result.get("Reason", ""))
                        row_data[f"{model_name} (Quote)"] = self._clean_text_for_excel(model_result.get("Quote", ""))
                    else:
                        # 如果没有找到结果，填充空值
                        if is_core_item:
                            row_data[f"{model_name} (Step 1)"] = ""
                            row_data[f"{model_name} (Step 2)"] = ""
                        else:
                            row_data[f"{model_name} (Step 1)"] = "NA"
                            row_data[f"{model_name} (Step 2)"] = ""
                        
                        row_data[f"{model_name} (Reason)"] = ""
                        row_data[f"{model_name} (Quote)"] = ""

                
                integrated_data.append(row_data)

        # 创建DataFrame
        if integrated_data:
            df_integrated = pd.DataFrame(integrated_data)
            
            # Create ExcelWriter object
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                # Save to "Integrated Results" sheet
                df_integrated.to_excel(writer, sheet_name="Integrated Results", index=False)
                
                # Get workbook and worksheet to apply formatting
                wb = writer.book
                ws = wb["Integrated Results"]
                
                # Apply highlighting format
                self._apply_highlighting(ws, df_integrated, model_names)
                
                # Adjust column widths
                self._adjust_column_widths(ws)
            
            logger.info(f"Integrated results have been successfully saved to {output_file}")
            
            # Summarize differences
            self._log_differences_summary(df_integrated, model_names)
        else:
            logger.warning("No data available to save to the integrated table")


    def _apply_highlighting(self, ws, df, model_names):
        """Apply highlighting based on discrepancies in Step 2 results"""
        # Define highlighting colors
        minor_diff_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
        major_diff_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        
        # Iterate through each row of data
        for row_idx, row_data in df.iterrows():
            excel_row_idx = row_idx + 2  # Excel row number (accounting for header row)
            
            # Collect Step 2 decisions for all models
            step2_decisions = {}
            for model_name in model_names:
                step2_col = f"{model_name} (Step 2)"
                if step2_col in row_data:
                    decision = row_data[step2_col]
                    if decision and decision != "":
                        step2_decisions[model_name] = decision
            
            # Check for discrepancies in Step 2
            if len(step2_decisions) > 1 and len(set(step2_decisions.values())) > 1:
                # Categorize decisions into low/high/unknown categories
                categorized_decisions = {}
                for model, decision in step2_decisions.items():
                    decision_lower = decision.lower() if isinstance(decision, str) else str(decision).lower()
                    if "low" in decision_lower:
                        if "definitely" in decision_lower:
                            category = "definitely_low"
                        else:
                            category = "probably_low"
                    elif "high" in decision_lower:
                        if "definitely" in decision_lower:
                            category = "definitely_high"
                        else:
                            category = "probably_high"
                    else:
                        category = "unknown"
                    categorized_decisions[model] = category
                
                # Check for major discrepancies (low vs high)
                categories = set(categorized_decisions.values())
                has_major_diff = (any("low" in c for c in categories) and any("high" in c for c in categories))
                has_minor_diff = len(categories) > 1 and not has_major_diff
                
                # Highlight entire row
                if has_major_diff or has_minor_diff:
                    fill_color = major_diff_fill if has_major_diff else minor_diff_fill
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=excel_row_idx, column=col_idx)
                        cell.fill = fill_color

    def _adjust_column_widths(self, ws):
        """Adjust column width"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 10)
            ws.column_dimensions[column].width = adjusted_width

    def _log_differences_summary(self, df, model_names):
        """Count and log difference summary"""
        diff_count = {"minor": 0, "major": 0, "total": 0}
        step1_diff_count = 0
        
        for _, row_data in df.iterrows():
            # 检查Step 2差异
            step2_decisions = {}
            for model_name in model_names:
                step2_col = f"{model_name} (Step 2)"
                if step2_col in row_data:
                    decision = row_data[step2_col]
                    if decision and decision != "":
                        step2_decisions[model_name] = decision
            
            if len(step2_decisions) > 1 and len(set(step2_decisions.values())) > 1:
                diff_count["total"] += 1
                
                # 分类差异类型
                categorized_decisions = {}
                for model, decision in step2_decisions.items():
                    decision_lower = decision.lower() if isinstance(decision, str) else str(decision).lower()
                    if "low" in decision_lower:
                        category = "low"
                    elif "high" in decision_lower:
                        category = "high"
                    else:
                        category = "unknown"
                    categorized_decisions[model] = category
                
                categories = set(categorized_decisions.values())
                if "low" in categories and "high" in categories:
                    diff_count["major"] += 1
                else:
                    diff_count["minor"] += 1
            
            # 检查Step 1差异（仅核心条目）
            step1_decisions = {}
            for model_name in model_names:
                step1_col = f"{model_name} (Step 1)"
                if step1_col in row_data:
                    decision = row_data[step1_col]
                    if decision and decision != "" and decision != "NA":
                        step1_decisions[model_name] = decision
            
            if len(step1_decisions) > 1 and len(set(step1_decisions.values())) > 1:
                step1_diff_count += 1
        
        logger.info(f"Total Step 2 assessment discrepancies: {diff_count['total']}")
        logger.info(f"  - Major discrepancies (low vs high): {diff_count['major']}")
        logger.info(f"  - Minor discrepancies (same type, different degrees): {diff_count['minor']}")
        logger.info(f"Total Step 1 assessment discrepancies: {step1_diff_count}")

    def detect_insufficient_evidence_patterns(self, reason_text):
        """Identify insufficient evidence based on patterns"""
        reason_lower = reason_text.lower()
        words = reason_lower.split()
        
        # 否定词
        negation_words = ['no', 'not', 'does not', 'did not', 'without', 'lacks', 'absence', 'didn', 'doesn']
        
        # 描述/报告类词汇
        description_words = ['explicit', 'clear', 'specific', 'detailed', 'direct', 'mention', 'state', 'describe', 'report', 'provide', 'information', 'details']
        
        # 检查3个词内的否定+描述组合
        for i, word in enumerate(words):
            if word in negation_words:
                # 检查后续3个词
                for j in range(i+1, min(i+4, len(words))):
                    if words[j] in description_words:
                        return True
        
        return False

    def _extract_text_with_retry(self, file_path: str, max_retries: int = 3) -> Optional[str]:
        """
        Extract text from document with retry logic
        
        Args:
            file_path: Path to the document
            max_retries: Maximum number of retry attempts
            
        Returns:
            Extracted text or None if failed
        """
        def extract_func():
            text = DocumentProcessor.extract_text(file_path)
            if not text:
                raise ROBError(
                    message=f"Empty text extracted from {file_path}",
                    category=ErrorCategory.DOCUMENT_PROCESSING_ERROR,
                    severity=ErrorSeverity.MEDIUM,
                    context={'file_path': file_path}
                )
            return text
        
        context = {'operation': 'text_extraction', 'file_path': file_path}
        result = self.error_handler.execute_with_retry(extract_func, context, max_retries)
        
        if result['success']:
            return result['result']
        else:
            logger.error(f"Failed to extract text from {file_path} after {result['attempts']} attempts")
            return None

    def _extract_study_id_with_retry(self, text: str, filename: str, max_retries: int = 3) -> str:
        """
        Extract study ID with retry logic
        
        Args:
            text: Document text
            filename: Document filename
            max_retries: Maximum number of retry attempts
            
        Returns:
            Extracted study ID
        """
        def extract_func():
            study_id = self.extract_study_id(text, filename, self.llm_configs[0])
            if not study_id or study_id == "Unknown, 0000":
                raise ROBError(
                    message=f"Invalid study ID extracted: {study_id}",
                    category=ErrorCategory.LLM_API_ERROR,
                    severity=ErrorSeverity.MEDIUM,
                    context={'filename': filename, 'study_id': study_id}
                )
            return study_id
        
        context = {'operation': 'study_id_extraction', 'filename': filename}
        result = self.error_handler.execute_with_retry(extract_func, context, max_retries)
        
        if result['success']:
            return result['result']
        else:
            logger.warning(f"Failed to extract valid study ID after {result['attempts']} attempts, using fallback")
            return self._extract_study_id_from_filename(filename)

    def _extract_study_id_from_filename(self, filename: str) -> str:
        """
        Extract study ID from filename as fallback
        
        Args:
            filename: Document filename
            
        Returns:
            Study ID extracted from filename or default
        """
        try:
            filename_match = re.search(r'([A-Z][a-zA-Z\s\'-]+).*?(\d{4})', filename)
            if filename_match:
                surname = filename_match.group(1).strip()
                year = filename_match.group(2)
                fallback_id = f"{surname}, {year}"
                logger.info(f"Extracted fallback study ID from filename: {fallback_id}")
                return fallback_id
        except Exception as e:
            logger.error(f"Failed to extract study ID from filename {filename}: {e}")
        
        return "Unknown, 0000"

    def _process_document_with_model_safe(self, file_path: str, content: Dict[str, Any], 
                                        llm_config: LLMConfig, model_idx: int, total_models: int) -> Dict[str, Any]:
        """
        Safely process document with a specific model, with comprehensive error handling
        
        Args:
            file_path: Path to the document file
            content: Document content dictionary
            llm_config: LLM configuration
            model_idx: Current model index
            total_models: Total number of models
            
        Returns:
            Dictionary containing processing result for this model
        """
        model_result = {
            'success': False,
            'errors': [],
            'core_results': [],
            'optional_results': [],
            'processing_time': 0
        }
        
        start_time = time.time()
        
        try:
            file_name = content['file_name']
            study_id = content['study_id']
            
            logger.info(f"Evaluating study {study_id} with model [{model_idx}/{total_models}]: {llm_config.name}")
            
            # Check if already processed
            checkpoint_key = f"{llm_config.name}:{file_path}"
            if checkpoint_key in self.processed_files:
                logger.info(f"Skipping already processed document: {study_id} (Model: {llm_config.name})")
                model_result['success'] = True
                return model_result
            
            # Process core items with error isolation
            core_success = self._process_core_items_safe(content, llm_config, model_result)
            
            # Process optional items if enabled and core succeeded
            optional_success = True
            if self.eval_optional_items:
                optional_success = self._process_optional_items_safe(content, llm_config, model_result)
            
            # Consider successful if core items processed successfully
            model_result['success'] = core_success
            
            if model_result['success']:
                # Save checkpoint
                self._save_checkpoint(file_path, llm_config.name)
                logger.info(f"Successfully processed {study_id} with model {llm_config.name}")
            else:
                logger.error(f"Failed to process {study_id} with model {llm_config.name}")
                
        except Exception as e:
            error_msg = f"Unexpected error processing with model {llm_config.name}: {str(e)}"
            model_result['errors'].append(error_msg)
            logger.error(error_msg)
        
        finally:
            model_result['processing_time'] = time.time() - start_time
        
        return model_result

    def _process_core_items_safe(self, content: Dict[str, Any], llm_config: LLMConfig, 
                               model_result: Dict[str, Any]) -> bool:
        """
        Safely process core items with retry logic
        
        Args:
            content: Document content dictionary
            llm_config: LLM configuration
            model_result: Model result dictionary to update
            
        Returns:
            True if successful, False otherwise
        """
        def process_func():
            study_id = content['study_id']
            file_name = content['file_name']
            text_snippet = content['text'][:self.max_text_length]
            base_file_name = os.path.splitext(file_name)[0]
            
            logger.info(f"Processing core items for {study_id}")
            
            core_prompt = self.generate_core_prompt(text_snippet, self.output_mode)
            
            if self.output_mode == 'json':
                if self.cost_analyzer:
                    core_response, usage_info = llm_config.generate_structured_response_with_usage(
                        core_prompt, 
                        response_format={"type": "json_schema", "json_schema": {"name": "core_assessment", "schema": CoreAssessmentResponse.model_json_schema()}}
                    )
                    if usage_info:
                        self.cost_analyzer.track_usage(
                            model=llm_config.model_name,
                            input_tokens=usage_info['input_tokens'],
                            output_tokens=usage_info['output_tokens'],
                            document_name=file_name,
                            operation="core_assessment_safe"
                        )
                else:
                    core_response = llm_config.generate_structured_response(
                        core_prompt, 
                        response_format={"type": "json_schema", "json_schema": {"name": "core_assessment", "schema": CoreAssessmentResponse.model_json_schema()}}
                    )
                core_results = self.parse_core_structured_response(study_id, core_response, base_file_name)
            elif self.output_mode == 'table':
                if self.cost_analyzer:
                    core_response, usage_info = llm_config.generate_response_with_usage(core_prompt)
                    if usage_info:
                        self.cost_analyzer.track_usage(
                            model=llm_config.model_name,
                            input_tokens=usage_info['input_tokens'],
                            output_tokens=usage_info['output_tokens'],
                            document_name=file_name,
                            operation="core_assessment_safe"
                        )
                else:
                    core_response = llm_config.generate_response(core_prompt)
                core_results = self.parse_core_table_response(study_id, core_response, base_file_name)
            else:
                raise ROBError(
                    message=f"Unsupported output mode: {self.output_mode}",
                    category=ErrorCategory.CONFIGURATION_ERROR,
                    severity=ErrorSeverity.HIGH,
                    context={'output_mode': self.output_mode}
                )
            
            if not core_results:
                raise ROBError(
                    message=f"Failed to parse core items response for {study_id}",
                    category=ErrorCategory.DATA_PARSING_ERROR,
                    severity=ErrorSeverity.MEDIUM,
                    context={'study_id': study_id, 'model': llm_config.name}
                )
            
            return core_results
        
        context = {
            'operation': 'core_items_processing',
            'study_id': content['study_id'],
            'model': llm_config.name
        }
        
        result = self.error_handler.execute_with_retry(process_func, context)
        
        if result['success']:
            core_results = result['result']
            self.results[llm_config.name]['core'].extend(core_results)
            model_result['core_results'] = core_results
            logger.info(f"Successfully processed core items for {content['study_id']}")
            return True
        else:
            # Log all errors from attempts
            for error_info in result.get('errors', []):
                model_result['errors'].append(error_info['error']['message'])
            
            logger.error(f"Failed to process core items for {content['study_id']} after {result['attempts']} attempts")
            return False

    def _process_optional_items_safe(self, content: Dict[str, Any], llm_config: LLMConfig, 
                                   model_result: Dict[str, Any]) -> bool:
        """
        Safely process optional items with retry logic
        
        Args:
            content: Document content dictionary
            llm_config: LLM configuration
            model_result: Model result dictionary to update
            
        Returns:
            True if successful, False otherwise
        """
        def process_func():
            study_id = content['study_id']
            file_name = content['file_name']
            text_snippet = content['text'][:self.max_text_length]
            base_file_name = os.path.splitext(file_name)[0]
            
            logger.info(f"Processing optional items for {study_id}")
            
            optional_prompt = self.generate_optional_prompt(text_snippet, self.output_mode)
            
            if self.output_mode == 'json':
                if self.cost_analyzer:
                    optional_response, usage_info = llm_config.generate_structured_response_with_usage(
                        optional_prompt,
                        response_format={"type": "json_schema", "json_schema": {"name": "optional_assessment", "schema": OptionalAssessmentResponse.model_json_schema()}}
                    )
                    if usage_info:
                        self.cost_analyzer.track_usage(
                            model=llm_config.model_name,
                            input_tokens=usage_info['input_tokens'],
                            output_tokens=usage_info['output_tokens'],
                            document_name=file_name,
                            operation="optional_assessment_safe"
                        )
                else:
                    optional_response = llm_config.generate_structured_response(
                        optional_prompt,
                        response_format={"type": "json_schema", "json_schema": {"name": "optional_assessment", "schema": OptionalAssessmentResponse.model_json_schema()}}
                    )
                optional_results = self.parse_optional_structured_response(study_id, optional_response, base_file_name)
            elif self.output_mode == 'table':
                if self.cost_analyzer:
                    optional_response, usage_info = llm_config.generate_response_with_usage(optional_prompt)
                    if usage_info:
                        self.cost_analyzer.track_usage(
                            model=llm_config.model_name,
                            input_tokens=usage_info['input_tokens'],
                            output_tokens=usage_info['output_tokens'],
                            document_name=file_name,
                            operation="optional_assessment_safe"
                        )
                else:
                    optional_response = llm_config.generate_response(optional_prompt)
                optional_results = self.parse_optional_table_response(study_id, optional_response, base_file_name)
            else:
                raise ROBError(
                    message=f"Unsupported output mode: {self.output_mode}",
                    category=ErrorCategory.CONFIGURATION_ERROR,
                    severity=ErrorSeverity.HIGH,
                    context={'output_mode': self.output_mode}
                )
            
            if not optional_results:
                raise ROBError(
                    message=f"Failed to parse optional items response for {study_id}",
                    category=ErrorCategory.DATA_PARSING_ERROR,
                    severity=ErrorSeverity.LOW,  # Low severity for optional items
                    context={'study_id': study_id, 'model': llm_config.name}
                )
            
            return optional_results
        
        context = {
            'operation': 'optional_items_processing',
            'study_id': content['study_id'],
            'model': llm_config.name
        }
        
        result = self.error_handler.execute_with_retry(process_func, context)
        
        if result['success']:
            optional_results = result['result']
            self.results[llm_config.name]['optional'].extend(optional_results)
            model_result['optional_results'] = optional_results
            logger.info(f"Successfully processed optional items for {content['study_id']}")
            return True
        else:
            # Log errors as warnings for optional items
            for error_info in result.get('errors', []):
                model_result['errors'].append(error_info['error']['message'])
            
            logger.warning(f"Failed to process optional items for {content['study_id']} after {result['attempts']} attempts")
            return False

