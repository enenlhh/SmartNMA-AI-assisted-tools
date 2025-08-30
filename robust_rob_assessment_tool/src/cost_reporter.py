"""
Cost Reporter for ROB Assessment Tool

This module provides comprehensive cost reporting functionality including
detailed reports, visualizations, and cost optimization recommendations.
"""

import os
import json
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd
from dataclasses import asdict

from cost_analyzer import CostAnalyzer, CostReport, ModelCostSummary


class CostReporter:
    """
    Comprehensive cost reporting system for ROB assessment tool
    
    Features:
    - Detailed cost reports with model breakdown
    - Cost optimization recommendations
    - Multi-format report generation (JSON, Excel, HTML)
    - Cost trend analysis and summaries
    """
    
    def __init__(self, cost_analyzer: CostAnalyzer):
        """
        Initialize cost reporter with cost analyzer
        
        Args:
            cost_analyzer: CostAnalyzer instance with usage data
        """
        self.cost_analyzer = cost_analyzer
    
    def generate_comprehensive_report(self, output_dir: str, 
                                    currencies: List[str] = None) -> Dict[str, str]:
        """
        Generate comprehensive cost report in multiple formats
        
        Args:
            output_dir: Directory to save reports
            currencies: List of currencies to include (default: ['USD', 'EUR', 'CNY'])
            
        Returns:
            Dictionary mapping format names to file paths
        """
        if currencies is None:
            currencies = ['USD', 'EUR', 'CNY']
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate base report data
        report_data = self._generate_report_data(currencies)
        
        # Save in different formats
        file_paths = {}
        
        # JSON report
        json_path = os.path.join(output_dir, f"cost_report_{self.cost_analyzer.session_id}.json")
        self._save_json_report(report_data, json_path)
        file_paths['json'] = json_path
        
        # Excel report
        try:
            excel_path = os.path.join(output_dir, f"cost_report_{self.cost_analyzer.session_id}.xlsx")
            self._save_excel_report(report_data, excel_path)
            file_paths['excel'] = excel_path
        except ImportError:
            print("Warning: openpyxl not available, skipping Excel report generation")
        
        # HTML report
        html_path = os.path.join(output_dir, f"cost_report_{self.cost_analyzer.session_id}.html")
        self._save_html_report(report_data, html_path)
        file_paths['html'] = html_path
        
        # Summary text report
        summary_path = os.path.join(output_dir, f"cost_summary_{self.cost_analyzer.session_id}.txt")
        self._save_summary_report(report_data, summary_path)
        file_paths['summary'] = summary_path
        
        return file_paths
    
    def _generate_report_data(self, currencies: List[str]) -> Dict[str, Any]:
        """
        Generate comprehensive report data structure
        
        Args:
            currencies: List of currencies to include
            
        Returns:
            Dictionary containing all report data
        """
        cost_summary = self.cost_analyzer.get_cost_summary()
        
        # Convert currency amounts
        currency_conversions = {}
        for currency in currencies:
            if currency != 'USD':
                try:
                    total_cost_converted = self.cost_analyzer.convert_currency(
                        cost_summary['total_cost_usd'], currency
                    )
                    currency_conversions[currency] = {
                        'total_cost': total_cost_converted,
                        'rate': self.cost_analyzer.pricing_config['currency_rates'][currency]
                    }
                except (KeyError, ValueError) as e:
                    print(f"Warning: Could not convert to {currency}: {e}")
        
        # Generate model breakdown with costs in multiple currencies
        model_breakdown = []
        for model_summary in cost_summary['model_summaries']:
            model_data = model_summary.copy()
            model_data['currency_costs'] = {'USD': model_summary['total_cost_usd']}
            
            for currency in currencies:
                if currency != 'USD' and currency in currency_conversions:
                    try:
                        converted_cost = self.cost_analyzer.convert_currency(
                            model_summary['total_cost_usd'], currency
                        )
                        model_data['currency_costs'][currency] = converted_cost
                    except (KeyError, ValueError):
                        pass
            
            model_breakdown.append(model_data)
        
        # Generate usage patterns analysis
        usage_patterns = self._analyze_usage_patterns()
        
        # Generate recommendations
        recommendations = self.cost_analyzer.generate_recommendations()
        
        # Calculate efficiency metrics
        efficiency_metrics = self._calculate_efficiency_metrics(cost_summary)
        
        report_data = {
            'metadata': {
                'session_id': self.cost_analyzer.session_id,
                'generated_at': datetime.now().isoformat(),
                'start_time': cost_summary['start_time'],
                'duration_minutes': cost_summary['duration_minutes'],
                'report_version': '1.0'
            },
            'summary': {
                'total_cost_usd': cost_summary['total_cost_usd'],
                'total_tokens': cost_summary['total_tokens'],
                'total_api_calls': cost_summary['total_api_calls'],
                'models_used': cost_summary['models_used'],
                'currency_conversions': currency_conversions
            },
            'model_breakdown': model_breakdown,
            'usage_patterns': usage_patterns,
            'efficiency_metrics': efficiency_metrics,
            'recommendations': recommendations,
            'detailed_usage': self._get_detailed_usage_data()
        }
        
        return report_data
    
    def _analyze_usage_patterns(self) -> Dict[str, Any]:
        """
        Analyze usage patterns from the cost analyzer data
        
        Returns:
            Dictionary containing usage pattern analysis
        """
        patterns = {
            'operations_breakdown': {},
            'documents_processed': set(),
            'peak_usage_periods': [],
            'model_preferences': {}
        }
        
        # Analyze operations breakdown
        operation_stats = {}
        document_count = 0
        
        for record in self.cost_analyzer.usage_records:
            # Track operations
            operation = record.operation or 'unknown'
            if operation not in operation_stats:
                operation_stats[operation] = {
                    'count': 0,
                    'total_tokens': 0,
                    'total_cost': 0.0
                }
            
            operation_stats[operation]['count'] += 1
            operation_stats[operation]['total_tokens'] += record.input_tokens + record.output_tokens
            
            # Calculate cost for this record
            input_cost, output_cost, total_cost = self.cost_analyzer.calculate_cost(
                record.model, record.input_tokens, record.output_tokens
            )
            operation_stats[operation]['total_cost'] += total_cost
            
            # Track documents
            if record.document_name:
                patterns['documents_processed'].add(record.document_name)
        
        patterns['operations_breakdown'] = operation_stats
        patterns['total_documents'] = len(patterns['documents_processed'])
        patterns['documents_processed'] = list(patterns['documents_processed'])  # Convert set to list for JSON
        
        # Analyze model usage preferences
        for model, usage in self.cost_analyzer.model_usage.items():
            patterns['model_preferences'][model] = {
                'api_calls': usage['api_calls'],
                'total_tokens': usage['input_tokens'] + usage['output_tokens'],
                'documents_processed': len(usage['documents'])
            }
        
        return patterns
    
    def _calculate_efficiency_metrics(self, cost_summary: Dict[str, Any]) -> Dict[str, Any]:
        """
        Calculate efficiency metrics for cost analysis
        
        Args:
            cost_summary: Cost summary data
            
        Returns:
            Dictionary containing efficiency metrics
        """
        metrics = {}
        
        if cost_summary['total_tokens'] > 0:
            metrics['cost_per_token'] = cost_summary['total_cost_usd'] / cost_summary['total_tokens']
        else:
            metrics['cost_per_token'] = 0.0
        
        if cost_summary['total_api_calls'] > 0:
            metrics['cost_per_api_call'] = cost_summary['total_cost_usd'] / cost_summary['total_api_calls']
            metrics['tokens_per_api_call'] = cost_summary['total_tokens'] / cost_summary['total_api_calls']
        else:
            metrics['cost_per_api_call'] = 0.0
            metrics['tokens_per_api_call'] = 0.0
        
        if cost_summary['duration_minutes'] > 0:
            metrics['cost_per_minute'] = cost_summary['total_cost_usd'] / cost_summary['duration_minutes']
            metrics['tokens_per_minute'] = cost_summary['total_tokens'] / cost_summary['duration_minutes']
        else:
            metrics['cost_per_minute'] = 0.0
            metrics['tokens_per_minute'] = 0.0
        
        # Calculate model efficiency comparison
        model_efficiency = {}
        for model_summary in cost_summary['model_summaries']:
            if model_summary['total_tokens'] > 0:
                efficiency = model_summary['total_cost_usd'] / model_summary['total_tokens']
                model_efficiency[model_summary['model']] = {
                    'cost_per_token': efficiency,
                    'relative_efficiency': 'N/A'  # Will be calculated after all models
                }
        
        # Calculate relative efficiency (compared to cheapest model)
        if model_efficiency:
            cheapest_cost_per_token = min(data['cost_per_token'] for data in model_efficiency.values())
            for model, data in model_efficiency.items():
                if cheapest_cost_per_token > 0:
                    relative_efficiency = data['cost_per_token'] / cheapest_cost_per_token
                    data['relative_efficiency'] = f"{relative_efficiency:.2f}x"
        
        metrics['model_efficiency'] = model_efficiency
        
        return metrics
    
    def _get_detailed_usage_data(self) -> List[Dict[str, Any]]:
        """
        Get detailed usage data for the report
        
        Returns:
            List of usage records with cost information
        """
        detailed_data = []
        
        for record in self.cost_analyzer.usage_records:
            input_cost, output_cost, total_cost = self.cost_analyzer.calculate_cost(
                record.model, record.input_tokens, record.output_tokens
            )
            
            detailed_data.append({
                'timestamp': record.timestamp.isoformat(),
                'model': record.model,
                'document_name': record.document_name,
                'operation': record.operation,
                'input_tokens': record.input_tokens,
                'output_tokens': record.output_tokens,
                'total_tokens': record.input_tokens + record.output_tokens,
                'input_cost_usd': input_cost,
                'output_cost_usd': output_cost,
                'total_cost_usd': total_cost
            })
        
        return detailed_data
    
    def _save_json_report(self, report_data: Dict[str, Any], file_path: str) -> None:
        """Save report as JSON file"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
    
    def _save_excel_report(self, report_data: Dict[str, Any], file_path: str) -> None:
        """Save report as Excel file with multiple sheets"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Add summary data
            summary_data = [
                ["Cost Analysis Summary", ""],
                ["Session ID", report_data['metadata']['session_id']],
                ["Generated At", report_data['metadata']['generated_at']],
                ["Duration (minutes)", report_data['metadata']['duration_minutes']],
                ["", ""],
                ["Total Cost (USD)", f"${report_data['summary']['total_cost_usd']:.4f}"],
                ["Total Tokens", report_data['summary']['total_tokens']],
                ["Total API Calls", report_data['summary']['total_api_calls']],
                ["Models Used", report_data['summary']['models_used']],
            ]
            
            for row_data in summary_data:
                ws_summary.append(row_data)
            
            # Format summary sheet
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Model breakdown sheet
            if report_data['model_breakdown']:
                ws_models = wb.create_sheet("Model Breakdown")
                
                # Flatten model data for Excel compatibility
                flattened_models = []
                for model in report_data['model_breakdown']:
                    flattened_model = {
                        'model': model['model'],
                        'api_calls': model['api_calls'],
                        'total_tokens': model['total_tokens'],
                        'total_input_tokens': model['total_input_tokens'],
                        'total_output_tokens': model['total_output_tokens'],
                        'input_cost_usd': model['input_cost_usd'],
                        'output_cost_usd': model['output_cost_usd'],
                        'total_cost_usd': model['total_cost_usd']
                    }
                    # Add currency costs as separate columns
                    if 'currency_costs' in model:
                        for currency, cost in model['currency_costs'].items():
                            flattened_model[f'cost_{currency.lower()}'] = cost
                    flattened_models.append(flattened_model)
                
                # Create DataFrame for model data
                model_df = pd.DataFrame(flattened_models)
                
                # Add headers
                for r in dataframe_to_rows(model_df, index=False, header=True):
                    ws_models.append(r)
                
                # Format headers
                for cell in ws_models[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Detailed usage sheet
            if report_data['detailed_usage']:
                ws_detailed = wb.create_sheet("Detailed Usage")
                
                # Create DataFrame for detailed usage
                detailed_df = pd.DataFrame(report_data['detailed_usage'])
                
                # Add headers
                for r in dataframe_to_rows(detailed_df, index=False, header=True):
                    ws_detailed.append(r)
                
                # Format headers
                for cell in ws_detailed[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            wb.save(file_path)
            
        except ImportError:
            raise ImportError("openpyxl is required for Excel report generation")
    
    def _save_html_report(self, report_data: Dict[str, Any], file_path: str) -> None:
        """Save report as HTML file"""
        html_content = self._generate_html_content(report_data)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _generate_html_content(self, report_data: Dict[str, Any]) -> str:
        """Generate HTML content for the report"""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ROB Assessment Cost Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        .header {{ background-color: #f0f0f0; padding: 20px; border-radius: 5px; }}
        .section {{ margin: 20px 0; }}
        .summary-table {{ border-collapse: collapse; width: 100%; }}
        .summary-table th, .summary-table td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        .summary-table th {{ background-color: #f2f2f2; }}
        .model-breakdown {{ margin: 20px 0; }}
        .recommendations {{ background-color: #fff3cd; padding: 15px; border-radius: 5px; }}
        .metric {{ display: inline-block; margin: 10px; padding: 10px; background-color: #e9ecef; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>ROB Assessment Cost Analysis Report</h1>
        <p><strong>Session ID:</strong> {report_data['metadata']['session_id']}</p>
        <p><strong>Generated:</strong> {report_data['metadata']['generated_at']}</p>
        <p><strong>Duration:</strong> {report_data['metadata']['duration_minutes']:.1f} minutes</p>
    </div>
    
    <div class="section">
        <h2>Cost Summary</h2>
        <div class="metric">
            <strong>Total Cost:</strong> ${report_data['summary']['total_cost_usd']:.4f} USD
        </div>
        <div class="metric">
            <strong>Total Tokens:</strong> {report_data['summary']['total_tokens']:,}
        </div>
        <div class="metric">
            <strong>API Calls:</strong> {report_data['summary']['total_api_calls']:,}
        </div>
        <div class="metric">
            <strong>Models Used:</strong> {report_data['summary']['models_used']}
        </div>
    </div>
    
    <div class="section">
        <h2>Model Breakdown</h2>
        <table class="summary-table">
            <thead>
                <tr>
                    <th>Model</th>
                    <th>API Calls</th>
                    <th>Total Tokens</th>
                    <th>Input Tokens</th>
                    <th>Output Tokens</th>
                    <th>Cost (USD)</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for model in report_data['model_breakdown']:
            html += f"""
                <tr>
                    <td>{model['model']}</td>
                    <td>{model['api_calls']}</td>
                    <td>{model['total_tokens']:,}</td>
                    <td>{model['total_input_tokens']:,}</td>
                    <td>{model['total_output_tokens']:,}</td>
                    <td>${model['total_cost_usd']:.4f}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
    </div>
    
    <div class="section">
        <h2>Efficiency Metrics</h2>
"""
        
        metrics = report_data['efficiency_metrics']
        html += f"""
        <div class="metric">
            <strong>Cost per Token:</strong> ${metrics['cost_per_token']:.6f}
        </div>
        <div class="metric">
            <strong>Cost per API Call:</strong> ${metrics['cost_per_api_call']:.4f}
        </div>
        <div class="metric">
            <strong>Tokens per API Call:</strong> {metrics['tokens_per_api_call']:.1f}
        </div>
        <div class="metric">
            <strong>Cost per Minute:</strong> ${metrics['cost_per_minute']:.4f}
        </div>
"""
        
        html += """
    </div>
    
    <div class="section recommendations">
        <h2>Cost Optimization Recommendations</h2>
        <ul>
"""
        
        for recommendation in report_data['recommendations']:
            html += f"<li>{recommendation}</li>"
        
        html += """
        </ul>
    </div>
    
    <div class="section">
        <h2>Usage Patterns</h2>
        <h3>Operations Breakdown</h3>
        <table class="summary-table">
            <thead>
                <tr>
                    <th>Operation</th>
                    <th>Count</th>
                    <th>Total Tokens</th>
                    <th>Total Cost (USD)</th>
                </tr>
            </thead>
            <tbody>
"""
        
        for operation, stats in report_data['usage_patterns']['operations_breakdown'].items():
            html += f"""
                <tr>
                    <td>{operation}</td>
                    <td>{stats['count']}</td>
                    <td>{stats['total_tokens']:,}</td>
                    <td>${stats['total_cost']:.4f}</td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""
        
        return html
    
    def _save_summary_report(self, report_data: Dict[str, Any], file_path: str) -> None:
        """Save a concise summary report as text file"""
        summary_lines = [
            "ROB ASSESSMENT COST ANALYSIS SUMMARY",
            "=" * 50,
            f"Session ID: {report_data['metadata']['session_id']}",
            f"Generated: {report_data['metadata']['generated_at']}",
            f"Duration: {report_data['metadata']['duration_minutes']:.1f} minutes",
            "",
            "COST SUMMARY",
            "-" * 20,
            f"Total Cost: ${report_data['summary']['total_cost_usd']:.4f} USD",
            f"Total Tokens: {report_data['summary']['total_tokens']:,}",
            f"Total API Calls: {report_data['summary']['total_api_calls']:,}",
            f"Models Used: {report_data['summary']['models_used']}",
            "",
            "MODEL BREAKDOWN",
            "-" * 20
        ]
        
        for model in report_data['model_breakdown']:
            summary_lines.extend([
                f"Model: {model['model']}",
                f"  API Calls: {model['api_calls']}",
                f"  Total Tokens: {model['total_tokens']:,}",
                f"  Cost: ${model['total_cost_usd']:.4f}",
                ""
            ])
        
        summary_lines.extend([
            "EFFICIENCY METRICS",
            "-" * 20,
            f"Cost per Token: ${report_data['efficiency_metrics']['cost_per_token']:.6f}",
            f"Cost per API Call: ${report_data['efficiency_metrics']['cost_per_api_call']:.4f}",
            f"Tokens per API Call: {report_data['efficiency_metrics']['tokens_per_api_call']:.1f}",
            "",
            "RECOMMENDATIONS",
            "-" * 20
        ])
        
        for i, recommendation in enumerate(report_data['recommendations'], 1):
            summary_lines.append(f"{i}. {recommendation}")
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(summary_lines))
    
    def generate_cost_comparison_report(self, other_sessions: List[str], 
                                      output_dir: str) -> Optional[str]:
        """
        Generate a cost comparison report across multiple sessions
        
        Args:
            other_sessions: List of other session IDs to compare with
            output_dir: Directory to save the comparison report
            
        Returns:
            Path to the comparison report file, or None if failed
        """
        # This would require loading other session data
        # Implementation would depend on how session data is stored
        # For now, return None as this is an advanced feature
        return None