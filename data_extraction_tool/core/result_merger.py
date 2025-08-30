#!/usr/bin/env python3
"""
结果合并模块
Result Merger Module

合并并行处理的提取结果
Merge extraction results from parallel processing
"""

import os
import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

try:
    from i18n.i18n_manager import get_message
except ImportError:
    def get_message(key, **kwargs):
        return key.format(**kwargs) if kwargs else key


class ResultMerger:
    """结果合并器 / Result Merger"""
    
    def __init__(self):
        self.merged_data = {}
        self.total_documents_processed = 0
        self.processing_stats = {}
    
    def merge_batch_results(self, temp_dir: str, config: Dict[str, Any]) -> str:
        """合并批次结果"""
        try:
            # 查找所有批次结果文件
            batch_files = list(Path(temp_dir).glob("batch_*_results.xlsx"))
            
            if not batch_files:
                raise Exception("No batch result files found")
            
            print(f"Found {len(batch_files)} batch result files")
            
            # 初始化合并数据结构
            self._initialize_merge_structure()
            
            # 逐个处理批次文件
            for batch_file in batch_files:
                self._merge_single_batch(batch_file)
            
            # 保存合并结果
            output_path = self._save_merged_results(config)
            
            # 生成处理统计报告
            self._generate_processing_report(output_path, config)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Failed to merge results: {str(e)}")
    
    def _initialize_merge_structure(self):
        """初始化合并数据结构"""
        # 基于数据提取的表结构初始化
        self.merged_data = {
            "Study_Info": [],
            "Groups": [],
            "Participant_Characteristics": [], 
            "Outcomes": [],
            "Results": [],
            "Comparisons": []
        }
        self.total_documents_processed = 0
        self.processing_stats = {
            "successful_extractions": 0,
            "failed_extractions": 0,
            "total_tables_extracted": 0,
            "processing_start_time": None,
            "processing_end_time": None
        }
    
    def _merge_single_batch(self, batch_file: Path):
        """合并单个批次文件"""
        try:
            # 读取Excel文件的所有工作表
            excel_data = pd.read_excel(batch_file, sheet_name=None)
            
            # 合并每个表的数据
            for sheet_name, df in excel_data.items():
                if sheet_name in self.merged_data:
                    if not df.empty:
                        # 将数据框转换为字典列表
                        records = df.to_dict('records')
                        self.merged_data[sheet_name].extend(records)
                        
                        # 更新统计信息
                        self.processing_stats["total_tables_extracted"] += 1
            
            # 更新文档处理计数
            if "Study_Info" in excel_data and not excel_data["Study_Info"].empty:
                self.total_documents_processed += len(excel_data["Study_Info"])
                self.processing_stats["successful_extractions"] += len(excel_data["Study_Info"])
            
            print(f"✓ Merged batch file: {batch_file.name}")
            
        except Exception as e:
            print(f"✗ Failed to merge batch file {batch_file.name}: {e}")
            self.processing_stats["failed_extractions"] += 1
    
    def _save_merged_results(self, config: Dict[str, Any]) -> str:
        """保存合并结果"""
        try:
            # 构建输出路径
            output_folder = config["paths"]["output_folder"]
            output_filename = config["paths"]["output_filename"]
            
            # 创建按日期组织的子目录结构
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            date_folder = datetime.now().strftime("%Y-%m-%d")
            
            # 创建结构化的输出目录
            structured_output_dir = os.path.join(output_folder, date_folder, "merged_results")
            Path(structured_output_dir).mkdir(parents=True, exist_ok=True)
            
            # 添加时间戳到文件名
            name_parts = output_filename.split('.')
            if len(name_parts) > 1:
                final_filename = f"{name_parts[0]}_merged_{timestamp}.{name_parts[1]}"
            else:
                final_filename = f"{output_filename}_merged_{timestamp}.xlsx"
            
            output_path = os.path.join(structured_output_dir, final_filename)
            
            # 创建Excel工作簿
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, data in self.merged_data.items():
                    if data:  # 只保存非空的表
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    else:
                        # 创建空的工作表以保持结构完整性
                        empty_df = pd.DataFrame()
                        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 美化Excel格式
            self._format_excel_output(output_path)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Failed to save merged results: {str(e)}")
    
    def _format_excel_output(self, output_path: str):
        """格式化Excel输出"""
        try:
            workbook = load_workbook(output_path)
            
            # 定义样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # 格式化标题行
                if worksheet.max_row > 0:
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                
                # 自动调整列宽
                from openpyxl.utils import get_column_letter
                for col_index, column in enumerate(worksheet.columns, 1):
                    max_length = 0
                    
                    # 使用get_column_letter函数安全地获取列字母
                    column_letter = get_column_letter(col_index)
                    
                    for cell in column:
                        try:
                            # 检查单元格是否为合并单元格
                            if hasattr(cell, 'value') and cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(output_path)
            
        except Exception as e:
            print(f"Warning: Failed to format Excel output: {e}")
    
    def _generate_processing_report(self, output_path: str, config: Dict[str, Any]):
        """生成处理统计报告"""
        try:
            report_data = {
                "总处理文档数": self.total_documents_processed,
                "成功提取": self.processing_stats["successful_extractions"],
                "失败提取": self.processing_stats["failed_extractions"],
                "提取的表格总数": self.processing_stats["total_tables_extracted"],
                "输出文件": output_path
            }
            
            # 计算各表的记录数
            for table_name, data in self.merged_data.items():
                report_data[f"{table_name}_记录数"] = len(data)
            
            # 将报告保存到同一目录下的reports子文件夹
            output_dir = os.path.dirname(output_path)
            reports_dir = os.path.join(os.path.dirname(output_dir), "reports")
            Path(reports_dir).mkdir(parents=True, exist_ok=True)
            
            report_filename = os.path.basename(output_path).replace('.xlsx', '_processing_report.json')
            report_path = os.path.join(reports_dir, report_filename)
            
            with open(report_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            
            # 打印摘要
            try:
                from i18n.i18n_manager import get_message
            except ImportError:
                def get_message(key, **kwargs):
                    fallback_messages = {
                        "processing_summary_title": "📋 Processing Summary",
                        "total_documents_processed": "Total Documents Processed: {count}",
                        "successful_extractions": "Successful Extractions: {count}",
                        "failed_extractions": "Failed Extractions: {count}",
                        "total_tables_extracted": "Total Tables Extracted: {count}",
                        "output_file": "Output File: {path}",
                        "report_file": "Report File: {path}"
                    }
                    message = fallback_messages.get(key, key)
                    return message.format(**kwargs) if kwargs and isinstance(message, str) else message
            
            print("\n" + "=" * 50)
            print(get_message("processing_summary_title"))
            print("=" * 50)
            print(get_message("total_documents_processed", count=self.total_documents_processed))
            print(get_message("successful_extractions", count=self.processing_stats['successful_extractions']))
            
            if self.processing_stats["failed_extractions"] > 0:
                print(get_message("failed_extractions", count=self.processing_stats['failed_extractions']))
            
            print(get_message("total_tables_extracted", count=self.processing_stats['total_tables_extracted']))
            
            # 显示各表记录数
            for table_name, data in self.merged_data.items():
                if data:
                    print(f"{table_name}: {len(data)} records")
            
            print(get_message("output_file", path=output_path))
            print(get_message("report_file", path=report_path))
            print("=" * 50)
            
        except Exception as e:
            print(f"Warning: Failed to generate processing report: {e}")
    
    def merge_results_from_state(self, state_file: str, output_dir: str = ".", 
                                final_output_prefix: str = "final_extraction_results",
                                backup_individual: bool = True) -> Dict[str, Any]:
        """从状态文件恢复并合并结果 / Merge results from state file"""
        try:
            print(f"🔄 Loading state from: {state_file}")
            
            with open(state_file, 'r', encoding='utf-8') as f:
                state = json.load(f)
            
            temp_dir = state.get("temp_dir")
            if not temp_dir or not os.path.exists(temp_dir):
                raise Exception(f"Temporary directory not found: {temp_dir}")
            
            print(f"📁 Found temporary directory: {temp_dir}")
            
            # 构建配置
            config = {
                "paths": {
                    "output_folder": output_dir,
                    "output_filename": f"{final_output_prefix}.xlsx"
                }
            }
            
            # 执行合并
            output_path = self.merge_batch_results(temp_dir, config)
            
            # 备份个别文件（如果需要）
            backup_info = {}
            if backup_individual:
                backup_info = self._backup_individual_files(temp_dir, output_dir)
            
            return {
                "excel_merge": {
                    "success": True,
                    "output_path": output_path,
                    "total_documents": self.total_documents_processed
                },
                "backup_info": backup_info,
                "processing_stats": self.processing_stats
            }
            
        except Exception as e:
            print(get_message("system_error", error=str(e)))
            return {
                "excel_merge": {
                    "success": False,
                    "error": str(e)
                }
            }
    
    def _backup_individual_files(self, temp_dir: str, output_dir: str) -> Dict[str, Any]:
        """备份个别批次文件 / Backup individual batch files"""
        try:
            import shutil
            from datetime import datetime
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            date_folder = datetime.now().strftime("%Y-%m-%d")
            
            # 创建结构化的备份目录
            backup_dir = os.path.join(output_dir, date_folder, "batch_results", f"batch_backups_{timestamp}")
            Path(backup_dir).mkdir(parents=True, exist_ok=True)
            
            batch_files = list(Path(temp_dir).glob("batch_*_results.xlsx"))
            backed_up_files = []
            
            for batch_file in batch_files:
                backup_path = os.path.join(backup_dir, batch_file.name)
                shutil.copy2(batch_file, backup_path)
                backed_up_files.append(backup_path)
            
            return {
                "backup_directory": backup_dir,
                "backed_up_files": backed_up_files,
                "file_count": len(backed_up_files)
            }
            
        except Exception as e:
            print(f"Warning: Failed to backup individual files: {e}")
            return {"error": str(e)}