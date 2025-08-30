"""
Full Text Extractor Module
Handles full-text screening workflow without prefiltering
"""

import openai
import os
import csv
import glob
import time
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

from .document_reader import DocumentReader

class FullTextExtractor:
    def __init__(self, screening_llm_configs, prompt_llm_config=None, positive_prompt_path=None, negative_prompt_path=None, config=None):
        """Initialize extractor with LLM configurations and system config"""
        self.screening_llm_configs = screening_llm_configs
        self.prompt_llm_config = prompt_llm_config or list(screening_llm_configs.values())[0]
        # Get first valid LLM name for prompt generation (skip comment fields)
        valid_llm_names = [name for name in screening_llm_configs.keys() if not name.startswith('_')]
        first_valid_name = valid_llm_names[0] if valid_llm_names else list(screening_llm_configs.keys())[0]
        self.prompt_llm_name = "Prompt LLM" if prompt_llm_config else first_valid_name
        
        self.positive_prompt_path = positive_prompt_path
        self.negative_prompt_path = negative_prompt_path
        
        # 读取配置参数
        self.config = config or {}
        self.mode = self.config.get('mode', {})
        self.parallel_settings = self.config.get('parallel_settings', {})
        self.resource_management = self.config.get('resource_management', {})
        
        # 并行处理配置
        self.screening_mode = self.mode.get('screening_mode', 'single')
        self.parallel_screeners = self.parallel_settings.get('parallel_screeners', 1)
        self.delay_between_screeners = self.resource_management.get('delay_between_screeners', 2)
        
        self.llm_clients = {}
        self.combined_prompt = None
        self.positive_prompt = None
        self.negative_prompt = None
        
        self.tokens_log = []
        self.tokens_csv_path = None
        
        # Initialize screening LLM clients
        for llm_name, config in self.screening_llm_configs.items():
            # Skip comment fields
            if llm_name.startswith('_'):
                continue
            client = openai.OpenAI(
                api_key=config['api_key'],
                base_url=config.get('base_url', 'https://vip.apiyi.com/v1')
            )
            self.llm_clients[llm_name] = client
        
        # Initialize prompt LLM client if separate
        if prompt_llm_config and self.prompt_llm_name not in self.llm_clients:
            prompt_client = openai.OpenAI(
                api_key=prompt_llm_config['api_key'],
                base_url=prompt_llm_config.get('base_url', 'https://vip.apiyi.com/v1')
            )
            self.llm_clients[self.prompt_llm_name] = prompt_client
    
    def create_combined_prompt(self, inclusion_criteria, exclusion_criteria=None, prompt_file_path=None):
        """Create combined prompt for full-text screening"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(f"\n🎯 {get_message('prompt_generation')}")
        
        # 基础prompt模板
        base_prompt = """You are a systematic review expert assisting with full text screening of research papers.
        Your task has two parts:
        
        PART 1: Extract the following PICOS information from the full text of a research paper:
        1. Study design - Be specific about the exact type of study. Only mark as randomized controlled trial if explicitly stated. Always extract the MOST SPECIFIC design mentioned.
        2. Participants - Include details of study subjects, demographics, conditions, severity, setting, etc.
        3. Intervention - Full description of the intervention being studied, including dosage, frequency, duration when available.
        4. Comparison - What the intervention is compared against, be specific about control groups.
        5. Outcomes - Primary and secondary outcomes measured in the study, including timepoints if mentioned.
        
        If any information is not explicitly mentioned, respond with "Not reported" for that category.
        If it is not possible to extract relevant information because a study is clearly irrelevant, respond with "Not available".
        """

        # 检查并加载现有提示词文件或生成新提示词
        if prompt_file_path and os.path.exists(prompt_file_path):
            print(f"✓ Found existing prompt file: {os.path.basename(prompt_file_path) if prompt_file_path else 'Unknown'}")
            print("Using existing screening guidance without regenerating")
            with open(prompt_file_path, "r", encoding='utf-8') as f:
                screening_guidance = f.read()
        else:
            print("Generating new full-text screening guidance...")
            screening_guidance = self._generate_fulltext_screening_guidance(inclusion_criteria, exclusion_criteria, prompt_file_path)
            if screening_guidance is None:
                print("⚠️ Falling back to basic screening format")
                return self._create_fallback_fulltext_prompt(base_prompt, inclusion_criteria, exclusion_criteria)

        # Task 2 prompt
        task2_prompt = f"""PART 2: Based on the extracted PICOS information, determine if the study meets ALL of the inclusion criteria:
        {screening_guidance}
        General Full-Text Screening Principles:
        - INCLUDE: If the full text clearly fulfills all inclusion criteria and no exclusion criteria
        - EXCLUDE: If the full text clearly violates any inclusion criterion or meets any exclusion criterion
        - EXCLUDE: If the full text is clearly not relevant to the inclusion criteria
        - UNCLEAR: Only if information is truly insufficient (rare in full-text, as complete information is available)
        - Be more definitive than title/abstract screening; use all available details for decision
        - Focus on conceptual matching rather than exact terminology
        - **Before finishing, MUST re-check all PICOS criteria systematically**
        Format your response EXACTLY as follows:
        
        ===PICOS EXTRACTION===
        Study design: [detailed extracted study design information]
        Participants: [detailed extracted participants information]
        Intervention: [detailed extracted intervention information]
        Comparison: [detailed extracted comparison information]
        Outcomes: [detailed extracted outcomes information]
        ===INCLUSION ASSESSMENT===
        Decision: [INCLUDE/EXCLUDE/UNCLEAR - choose only one]
        Explanation: [brief rationale in ONE SENTENCE, 50 WORDS MAXIMUM, focus on PRIMARY reason]
        """
        
        complete_prompt = base_prompt + task2_prompt
        final_template_tokens = len(complete_prompt) // 4
        print(f"✓ Prompt template ready (~{final_template_tokens:,} tokens)")
        return complete_prompt

    def _generate_fulltext_screening_guidance(self, inclusion_criteria, exclusion_criteria, prompt_file_path):
        """Generate new full-text screening guidance"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print("📋 Processing criteria and generating detailed screening guidance...")
        
        # 准备标准文本
        print("\n🔍 Analyzing inclusion/exclusion criteria:")
        inclusion_details = {}
        exclusion_details = {}
        
        for key, value in inclusion_criteria.items():
            if value and value not in ["No limit", "None"]:
                inclusion_details[key] = value
                print(f"  ✓ Include - {key}: {value[:80]}{'...' if len(value) > 80 else ''}")
        
        if exclusion_criteria:
            for key, value in exclusion_criteria.items():
                if value and value not in ["No limit", "None"]:
                    exclusion_details[key] = value
                    print(f"  ✗ Exclude - {key}: {value[:80]}{'...' if len(value) > 80 else ''}")

        # 为全文筛选定制的prompt生成指令
        screening_prompt = f"""As a systematic review expert, create COMPREHENSIVE full-text screening guidance based on these criteria:

        INCLUSION CRITERIA:
        {chr(10).join(f"{k}: {v}" for k, v in inclusion_details.items())}
        
        {"EXCLUSION CRITERIA:" if exclusion_details else ""}
        {chr(10).join(f"{k}: {v}" for k, v in exclusion_details.items()) if exclusion_details else ""}
        
        Create guidance that:
        1. Provides direct instructions for each PICOS criterion
        2. Emphasizes definitive decisions using full text's complete information
        3. Converts any indirect/ambiguous criteria into explicit, actionable statements
        4. Keeps it brief and professional, trusting screener's knowledge
        
        FORMAT YOUR GUIDANCE LIKE THIS:
        
        STUDY DESIGN SCREENING:
        Must Include: [Detailed description of acceptable study designs based on criteria]
        Must Exclude: [Detailed description of study designs that should be excluded]
        Decision Rules: [Specific rules for making inclusion decisions based on study design]
        
        PARTICIPANTS SCREENING:
        Must Include: [Detailed description of target population characteristics]
        Must Exclude: [Detailed description of populations that should be excluded]
        Decision Rules: [Specific rules for evaluating participant eligibility]
        
        INTERVENTION SCREENING:
        Must Include: [Detailed description of target interventions]
        Must Exclude: [Detailed description of interventions that should be excluded]
        Decision Rules: [Specific rules for evaluating intervention relevance]
        
        COMPARISON SCREENING:
        Must Include: [Detailed description of acceptable comparisons]
        Must Exclude: [Detailed description of comparisons that should be excluded]
        Decision Rules: [Specific rules for evaluating comparison groups]
        
        OUTCOMES SCREENING:
        Must Include: [Detailed description of target outcomes]
        Must Exclude: [Detailed description of outcomes that should be excluded]
        Decision Rules: [Specific rules for evaluating outcome relevance]
        
        OVERALL DECISION FRAMEWORK:
        - INCLUDE: Study meets ALL inclusion criteria and NO exclusion criteria
        - EXCLUDE: Study fails to meet ANY inclusion criterion OR meets ANY exclusion criterion
        - UNCLEAR: Insufficient information to make definitive decision (rare in full-text screening)
        
        IMPORTANT NOTES:
        - Full-text screening allows for more definitive decisions than title/abstract screening
        - Use the complete study information available in the full text
        - Be thorough but decisive - avoid UNCLEAR unless truly ambiguous
        - Focus on conceptual matching rather than exact terminology
        - Consider the study's primary focus and main findings
        
        Your guidance will be used for comprehensive full-text evaluation of research papers.
        """

        # 估算prompt tokens
        prompt_tokens_estimate = len(screening_prompt) // 4
        print(f"📊 Estimated tokens for guidance generation: ~{prompt_tokens_estimate:,}")
        
        max_retries = 5
        for retry_count in range(max_retries):
            try:
                print(f"🤖 Sending request to {self.prompt_llm_name} for guidance generation...")
                client = self.llm_clients[self.prompt_llm_name]
                config = self.prompt_llm_config if self.prompt_llm_name == "Prompt LLM" else self.screening_llm_configs[self.prompt_llm_name]
                
                response = client.chat.completions.create(
                    model=config['model'],
                    messages=[
                        {"role": "system", "content": "You are an evidence-based medicine expert specializing in systematic reviews and full-text screening."},
                        {"role": "user", "content": screening_prompt}
                    ],
                    temperature=0.1,
                    max_tokens=4000  # 增加token限制以适应更详细的全文指导
                )
                
                screening_guidance = response.choices[0].message.content
                
                try:
                    prompt_tokens = response.usage.prompt_tokens
                    completion_tokens = response.usage.completion_tokens
                    total_tokens = response.usage.total_tokens
                    print(f"✓ Tokens used: {total_tokens:,} (Prompt: {prompt_tokens:,}, Completion: {completion_tokens:,})")
                except AttributeError:
                    guidance_tokens_estimate = len(screening_guidance) // 4
                    print(f"✓ Estimated tokens for generated guidance: ~{guidance_tokens_estimate:,}")
                
                print("✅ Full-text screening guidance successfully generated!")
                if prompt_file_path:
                    with open(prompt_file_path, "w", encoding='utf-8') as f:
                        f.write(screening_guidance)
                    
                    print(f"\n" + "="*60)
                    print("🎉 FULL-TEXT SCREENING GUIDANCE GENERATED")
                    print("="*60)
                    print("The full-text screening guidance has been successfully generated!")
                    print(f"\n📁 Generated File:")
                    print(f"   • Full-Text Screening Guidance: {prompt_file_path}")
                    
                    print("\n📋 Next Steps:")
                    print("1. Please review and adjust the generated full-text screening guidance as needed")
                    print("2. The guidance contains detailed instructions for comprehensive full-text evaluation")
                    print("3. After reviewing, restart the program to begin full-text screening process")
                    
                    print(f"\n🔄 To restart: python run_fulltext.py")
                    print("="*60)
                    
                    import sys
                    sys.exit(0)
                else:
                    return screening_guidance
                
            except Exception as e:
                if retry_count < max_retries - 1:
                    print(f"Error generating full-text guidance, attempt {retry_count + 1}/{max_retries}: {str(e)}. Retrying...")
                    time.sleep(2 ** retry_count)
                else:
                    print(f"\nERROR: Failed to generate full-text screening guidance after {max_retries} attempts: {str(e)}")
                    return None

    def _create_fallback_fulltext_prompt(self, base_prompt, inclusion_criteria, exclusion_criteria):
        """Create fallback prompt for full-text screening"""
        print("Falling back to basic full-text screening format...")
        
        inclusion_details = {k: v for k, v in inclusion_criteria.items() if v and v not in ["No limit", "None"]}
        exclusion_details = {k: v for k, v in exclusion_criteria.items() if v and v not in ["No limit", "None"]} if exclusion_criteria else {}
        
        # 简单的fallback方案，专门针对全文筛选
        fallback_guidance = "### Full-Text Screening Guidance\n\n"
        fallback_guidance += "**Inclusion Criteria (All must be met):**\n"
        for key, value in inclusion_details.items():
            fallback_guidance += f"- {key}: {value}\n  * Evaluate based on complete study information in full text\n  * Look for explicit statements and detailed descriptions\n\n"
        
        if exclusion_details:
            fallback_guidance += "\n**Exclusion Criteria (Any one excludes the study):**\n"
            for key, value in exclusion_details.items():
                fallback_guidance += f"- {key}: {value}\n  * Exclude if clearly meeting this criterion based on full text\n\n"
        
        fallback_guidance += "\n**Full-Text Screening Principles:**\n"
        fallback_guidance += "- Use all available information in the complete document\n"
        fallback_guidance += "- Be more definitive than title/abstract screening\n"
        fallback_guidance += "- INCLUDE only if ALL inclusion criteria are met\n"
        fallback_guidance += "- EXCLUDE if ANY exclusion criterion is met or ANY inclusion criterion is not met\n"
        fallback_guidance += "- UNCLEAR should be rare since full text provides complete information\n"
        fallback_guidance += "- Focus on the study's main objectives and primary findings\n"
        
        # 添加到Task 2 prompt
        task2_prompt = f"""PART 2: Based on the extracted PICOS information, determine if the study meets ALL inclusion criteria:

        {fallback_guidance}

        Format your response EXACTLY as follows:
        
        <START_PICOS>
        <STUDY_DESIGN>detailed extracted study design information</STUDY_DESIGN>
        <PARTICIPANTS>detailed extracted participants information</PARTICIPANTS>
        <INTERVENTION>detailed extracted intervention information</INTERVENTION>
        <COMPARISON>detailed extracted comparison information</COMPARISON>
        <OUTCOMES>detailed extracted outcomes information</OUTCOMES>
        </START_PICOS>

        <DECISION>INCLUDE or EXCLUDE or UNCLEAR - choose only one of these three options</DECISION>

        <EXPLANATION>
        detailed explanation for your decision, citing specific criteria that were met or not met
        </EXPLANATION>
        """
        
        complete_prompt = base_prompt + task2_prompt
        fallback_template_tokens = len(complete_prompt) // 4
        print(f"\nEstimated tokens for fallback full-text prompt template: ~{fallback_template_tokens} tokens")
        print("Using fallback full-text screening guidance due to LLM error.")
        return complete_prompt

    
    def load_or_generate_positive_negative_prompts(self, normal_prompt):
        """Load or generate positive and negative versions of prompt for full-text screening"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(f"\n🔍 {get_message('positive_negative_check')}")
        
        # Check if files exist
        positive_exists = self.positive_prompt_path and os.path.exists(self.positive_prompt_path)
        negative_exists = self.negative_prompt_path and os.path.exists(self.negative_prompt_path)
        
        if positive_exists and negative_exists:
            print(f"✓ Found positive prompt: {os.path.basename(self.positive_prompt_path) if self.positive_prompt_path else 'Unknown'}")
            print(f"✓ Found negative prompt: {os.path.basename(self.negative_prompt_path) if self.negative_prompt_path else 'Unknown'}")
            print(get_message("variants_found"))
            
            # Load from files directly
            if self.positive_prompt_path and self.negative_prompt_path:
                with open(self.positive_prompt_path, "r", encoding='utf-8') as f:
                    positive_prompt = f.read()
                with open(self.negative_prompt_path, "r", encoding='utf-8') as f:
                    negative_prompt = f.read()
            else:
                print("⚠️ Prompt file paths not properly set")
                return self.generate_positive_negative_fulltext_prompts(normal_prompt)
                
            print("✅ Successfully loaded existing prompt variants")
            return positive_prompt, negative_prompt
        
        else:
            print(get_message("generating_variants"))
            return self.generate_positive_negative_fulltext_prompts(normal_prompt)
            
    def generate_positive_negative_fulltext_prompts(self, normal_prompt):
        """Generate positive and negative versions of full-text screening prompt"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print("🔄 Generating positive and negative prompt variants...")
        
        # 先检查是否已有有效的提示文件
        if (self.positive_prompt_path and os.path.exists(self.positive_prompt_path) and 
            self.negative_prompt_path and os.path.exists(self.negative_prompt_path)):
            
            if self.positive_prompt_path and self.negative_prompt_path:
                with open(self.positive_prompt_path, "r", encoding='utf-8') as f:
                    positive_content = f.read().strip()
                with open(self.negative_prompt_path, "r", encoding='utf-8') as f:
                    negative_content = f.read().strip()
            else:
                # 如果路径为 None，跳过文件检查直接生成新的
                positive_content = ""
                negative_content = ""
                
            # 检查文件是否包含实际内容（不是占位符）
            if (positive_content and negative_content and 
                "[restructured full-text prompt" not in positive_content and 
                "[restructured full-text prompt" not in negative_content and
                len(positive_content) > 500 and len(negative_content) > 500):  # 全文prompt更长
                print("✓ Found valid existing prompt variant files")
                return positive_content, negative_content
        # 如果文件不存在或无效，生成新的全文筛选提示
        prompt_generation_instruction = f"""
        Based on the following original full-text screening prompt, create TWO versions with subtly different presentation styles while preserving ALL information content exactly:
        
        ORIGINAL FULL-TEXT PROMPT:
        {normal_prompt}
        
        Create these two versions specifically optimized for full-text screening:
        
        1. VERSION A (Comprehensive Full-Text Evaluation): 
        - Present criteria with emphasis on thorough, systematic evaluation of complete document
        - Use language that encourages comprehensive assessment of all available information
        - Structure information to promote detailed examination of full study content
        - Emphasize the advantage of having complete study information for decision-making
        
        2. VERSION B (Precision-Focused Full-Text Analysis):
        - Present criteria with emphasis on precise, targeted evaluation of key study elements
        - Use language that emphasizes meeting exact requirements based on full study details
        - Structure information to encourage focused verification of specific criteria
        - Emphasize accuracy and specificity in decision-making with complete information
        
        CRITICAL REQUIREMENTS FOR FULL-TEXT SCREENING:
        - Do NOT add, remove, or modify ANY factual content or criteria
        - Do NOT change any specific requirements or thresholds
        - Do NOT use obvious directional language like "lean toward inclusion/exclusion"
        - ONLY change the presentation flow and subtle linguistic framing
        - Preserve all PICOS extraction instructions exactly
        - Maintain the same response format requirements (===PICOS EXTRACTION=== ... ===INCLUSION ASSESSMENT=== Decision: ... Explanation: ...)
        - Keep all technical details identical
        - Use neutral, professional tone throughout
        - Account for the fact that full-text provides complete study information
        - Maintain the higher decisiveness expected in full-text vs title/abstract screening
        
        The two versions should feel naturally different in their approach to full-text evaluation but not obviously biased toward any particular outcome.
        
        Please format your response as:
        ===VERSION A (COMPREHENSIVE FULL-TEXT)===
        [restructured full-text prompt with comprehensive evaluation approach]
        
        ===VERSION B (PRECISION-FOCUSED FULL-TEXT)===
        [restructured full-text prompt with precision-focused approach]
        """
                
        max_retries = 3
        for retry_count in range(max_retries):
            try:
                client = self.llm_clients[self.prompt_llm_name]
                config = self.prompt_llm_config if self.prompt_llm_name == "Prompt LLM" else self.screening_llm_configs[self.prompt_llm_name]
                
                response = client.chat.completions.create(
                    model=config['model'],
                    messages=[
                        {"role": "system", "content": "You are an expert at restructuring full-text screening content with different psychological framing while preserving all factual information exactly."},
                        {"role": "user", "content": prompt_generation_instruction}
                    ],
                    temperature=0.1,
                    max_tokens=6000  # 增加token限制，因为全文prompt更长
                )
                
                result = response.choices[0].message.content
                
                # 改进的解析逻辑，适配全文筛选
                if "===VERSION A (COMPREHENSIVE FULL-TEXT)===" in result and "===VERSION B (PRECISION-FOCUSED FULL-TEXT)===" in result:
                    try:
                        # 分割响应
                        parts = result.split("===VERSION B (PRECISION-FOCUSED FULL-TEXT)===")
                        if len(parts) >= 2:
                            positive_prompt = parts[0].replace("===VERSION A (COMPREHENSIVE FULL-TEXT)===", "").strip()
                            negative_prompt = parts[1].strip()
                            
                            # 验证内容不为空且不是占位符
                            if (positive_prompt and negative_prompt and 
                                len(positive_prompt) > 500 and len(negative_prompt) > 500 and
                                "[restructured full-text prompt" not in positive_prompt and 
                                "[restructured full-text prompt" not in negative_prompt):
                                
                                # 保存到文件
                                if self.positive_prompt_path:
                                    with open(self.positive_prompt_path, "w", encoding='utf-8') as f:
                                        f.write(positive_prompt)
                                    print(f"✓ Positive full-text prompt (comprehensive evaluation) saved to: {self.positive_prompt_path}")
                                
                                if self.negative_prompt_path:
                                    with open(self.negative_prompt_path, "w", encoding='utf-8') as f:
                                        f.write(negative_prompt)
                                    print(f"✓ Negative full-text prompt (precision-focused) saved to: {self.negative_prompt_path}")
                                
                                print("✓ Successfully generated and saved Positive and Negative full-text prompt versions")
                                
                                # 直接退出程序，提示用户检查文件
                                print(f"\n" + "="*60)
                                print("🎉 FULL-TEXT PROMPT GENERATION COMPLETED")
                                print("="*60)
                                print("Positive and Negative full-text prompt variants have been successfully generated!")
                                print("\n📁 Generated Files:")
                                if self.positive_prompt_path:
                                    print(f"   • Positive Full-Text Prompt (Comprehensive): {self.positive_prompt_path}")
                                if self.negative_prompt_path:
                                    print(f"   • Negative Full-Text Prompt (Precision-Focused): {self.negative_prompt_path}")
                                
                                print("\n📋 Next Steps:")
                                print("1. Please review and adjust the generated full-text prompt files as needed")
                                print("2. The prompts use different psychological framing for full-text evaluation")
                                print("3. After reviewing, restart the program to begin full-text screening process")
                                print("4. The system will automatically use the generated full-text prompt variants")
                                
                                print(f"\n🔄 To restart: python run_fulltext.py")
                                print("="*60)
                                
                                # 直接退出程序
                                import sys
                                sys.exit(0)
                            else:
                                raise ValueError("Generated full-text prompts are too short, contain placeholders, or don't have the expected structure")
                        else:
                            raise ValueError("Could not split full-text response properly")
                    except Exception as parse_error:
                        print(f"Parsing error: {parse_error}")
                        raise ValueError(f"Failed to parse generated full-text prompt versions: {parse_error}")
                else:
                    raise ValueError("Response does not contain expected section markers for full-text prompts")
                    
            except Exception as e:
                if retry_count < max_retries - 1:
                    print(f"Failed to generate full-text prompt variants, retry {retry_count + 1}/{max_retries}: {str(e)}")
                    time.sleep(2)
                else:
                    print(f"Failed to generate full-text prompt variants after {max_retries} attempts: {str(e)}")
                    print("Using original full-text prompt for both positive and negative versions")
                    
                    # 使用原始提示作为回退
                    if self.positive_prompt_path:
                        with open(self.positive_prompt_path, "w", encoding='utf-8') as f:
                            f.write(normal_prompt)
                        print(f"✓ Fallback: Original full-text prompt saved as positive prompt to: {self.positive_prompt_path}")
                    
                    if self.negative_prompt_path:
                        with open(self.negative_prompt_path, "w", encoding='utf-8') as f:
                            f.write(normal_prompt)
                        print(f"✓ Fallback: Original full-text prompt saved as negative prompt to: {self.negative_prompt_path}")
                    
                    # 回退情况也直接退出
                    print(f"\n" + "="*60)
                    print("⚠️  FULL-TEXT PROMPT GENERATION FALLBACK")
                    print("="*60)
                    print("Due to generation issues, the original full-text prompt has been saved to both files.")
                    print("\n📁 Fallback Files:")
                    if self.positive_prompt_path:
                        print(f"   • Positive Full-Text Prompt: {self.positive_prompt_path}")
                    if self.negative_prompt_path:
                        print(f"   • Negative Full-Text Prompt: {self.negative_prompt_path}")
                    
                    print("\n📋 Next Steps:")
                    print("1. Please manually edit the full-text prompt files to create different variants")
                    print("2. Adjust the psychological framing for full-text screening as needed")
                    print("3. After editing, restart the program to begin full-text screening process")
                    
                    print(f"\n🔄 To restart: python run_fulltext.py")
                    print("="*60)
                    
                    import sys
                    sys.exit(0)
        
        return normal_prompt, normal_prompt

    def process_study_sequential(self, document_content, inclusion_criteria, study_id):
        """Sequential processing of one full-text study: LLM A processes first, then LLM B uses appropriate prompt based on result"""
        print("  Starting sequential processing of one full-text study...")
        
        # Ensure we have the combined prompt
        if self.combined_prompt is None:
            self.combined_prompt = self.create_combined_prompt(inclusion_criteria)
        
        # Load or generate positive and negative versions of prompt
        if self.positive_prompt is None or self.negative_prompt is None:
            self.positive_prompt, self.negative_prompt = self.load_or_generate_positive_negative_prompts(self.combined_prompt)
        
        # Get LLM names list (filter out comment fields)
        all_llm_names = list(self.screening_llm_configs.keys())
        llm_names = [name for name in all_llm_names if not name.startswith('_')]
        if len(llm_names) < 2:
            raise ValueError(f"At least 2 LLM instances required for sequential processing, found {len(llm_names)}: {llm_names}")
        
        llm_a_name = llm_names[0]  # First LLM
        llm_b_name = llm_names[1]  # Second LLM
        
        results = {}
        
        # Step 1: LLM A processes with normal prompt
        print(f"  Step 1: {llm_a_name} processing full-text with normal prompt...")
        try:
            picos_a, inclusion_a = self.process_fulltext_with_prompt(
                llm_a_name, document_content, self.combined_prompt, study_id
            )
            results[llm_a_name] = {
                'picos': picos_a,
                'inclusion': inclusion_a
            }
            print(f"  {llm_a_name} completed: {inclusion_a}")
            
            # Determine LLM A's decision
            if "INCLUDE" in inclusion_a:
                llm_a_decision = "INCLUDE"
            elif "EXCLUDE" in inclusion_a:
                llm_a_decision = "EXCLUDE"
            else:
                llm_a_decision = "UNCLEAR"
            
            # Step 2: Select prompt for LLM B based on LLM A's result
            if llm_a_decision == "INCLUDE":
                selected_prompt = self.negative_prompt
                prompt_type = "negative prompt (precision-focused evaluation)"
                print(f"  LLM A result is INCLUDE, LLM B will use {prompt_type}")
            elif llm_a_decision == "EXCLUDE":
                selected_prompt = self.positive_prompt  
                prompt_type = "positive prompt (comprehensive evaluation)"
                print(f"  LLM A result is EXCLUDE, LLM B will use {prompt_type}")
            else:
                # Use positive prompt for UNCLEAR cases
                selected_prompt = self.positive_prompt
                prompt_type = "positive prompt (comprehensive evaluation, because LLM A result is unclear)"
                print(f"  LLM A result is UNCLEAR, LLM B will use {prompt_type}")
            
            # Step 3: LLM B processes with selected prompt
            print(f"  Step 2: {llm_b_name} processing full-text with {prompt_type}...")
            picos_b, inclusion_b = self.process_fulltext_with_prompt(
                llm_b_name, document_content, selected_prompt, study_id
            )
            results[llm_b_name] = {
                'picos': picos_b,
                'inclusion': inclusion_b,
                'prompt_type': prompt_type
            }
            print(f"  {llm_b_name} completed: {inclusion_b}")
            
        except Exception as e:
            print(f"  Sequential full-text processing error: {str(e)}")
            # If error occurs, return error information
            error_result = {
                'picos': {
                    "study_design": f"Error: {str(e)}",
                    "participants": f"Error: {str(e)}",
                    "intervention": f"Error: {str(e)}",
                    "comparison": f"Error: {str(e)}",
                    "outcomes": f"Error: {str(e)}"
                },
                'inclusion': f"UNCLEAR (Error: {str(e)})"
            }
            results[llm_a_name] = error_result
            results[llm_b_name] = error_result
        
        print(f"  Sequential full-text processing completed")
        return results

    def extract_study_id(self, document_text, llm_name):
        """Extract study ID (author surname and year)"""
        client = self.llm_clients[llm_name]
        config = self.screening_llm_configs[llm_name]
        
        prompt = f"""
        Extract from the text:
        1. The surname of the first author
        2. The publication year (4 digits), if you can't find the year, use the newest year you can find in the text
        Return format: "Surname, YYYY"
        If not found, return: "Unknown, 0000"
        Text:
        {document_text[:1500]}
        """
        
        max_retries = 3
        for retry_count in range(max_retries):
            try:
                response = client.chat.completions.create(
                    model=config['model'],
                    messages=[
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.0,
                )
                
                result = response.choices[0].message.content.strip()
                
                # 1. 首先尝试使用原始正则表达式
                if re.match(r"^[A-Za-z\-'\s]+,?\s*\d{4}$", result):
                    return result
                
                # 2. 如果原始正则表达式失败，尝试一个更宽松的正则表达式
                match = re.search(r"([A-Za-z\-'\s]+)[\s,]+(\d{4})", result)
                if match:
                    surname = match.group(1).strip()
                    year = match.group(2)
                    formatted_result = f"{surname}, {year}"
                    return formatted_result
                
                # 3. 尝试查找任何姓氏和四位数字的组合
                name_match = re.search(r"([A-Za-z\-'\s]+)", result)
                year_match = re.search(r"(\d{4})", result)
                if name_match and year_match:
                    surname = name_match.group(1).strip()
                    year = year_match.group(1)
                    formatted_result = f"{surname}, {year}"
                    return formatted_result
                
                return "Unknown, 0000"
                
            except Exception as e:
                if retry_count < max_retries - 1:
                    print(f"  Error extracting study ID: {str(e)}. Retrying {retry_count+1}/{max_retries}...")
                    time.sleep(2)
                else:
                    print(f"  Failed to extract study ID after {max_retries} attempts: {str(e)}")
                    return "Unknown, 0000"
    
    def process_single_document(self, file_path, inclusion_criteria, file_index, total_files):
        """处理单个文档（用于并行处理）"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            # Fallback if i18n not available
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        file_name = os.path.basename(file_path)
        
        try:
            # 获取文件大小
            try:
                file_size_kb = os.path.getsize(file_path) / 1024
            except:
                file_size_kb = 0
            
            # 读取文档内容
            document_content = DocumentReader.read_file(file_path)
            if document_content is None:
                return None
            
            # 显示提取的文本信息
            char_count = len(document_content)
            
            # 提取研究ID
            valid_llm_names = [name for name in self.screening_llm_configs.keys() if not name.startswith('_')]
            llm_name = valid_llm_names[0] if valid_llm_names else list(self.screening_llm_configs.keys())[0]
            study_id = self.extract_study_id(document_content, llm_name)
            
            # 处理文档（无预筛选）
            start_time = time.time()
            study_results = self.process_study_sequential(document_content, inclusion_criteria, study_id)
            end_time = time.time()
            
            # 准备结果数据
            results = []
            for llm_name, result in study_results.items():
                result_data = {
                    'filename': file_name,
                    'filepath': file_path,
                    'study_id': study_id,
                    'llm_name': llm_name,
                    'picos': result['picos'],
                    'inclusion': result['inclusion']
                }
                results.append(result_data)
            
            return results
            
        except Exception as e:
            return None
    
    def process_documents_parallel(self, folder_path, inclusion_criteria, output_path, exclusion_criteria=None, prompt_file_path=None):
        """并行处理所有文档"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            # Fallback if i18n not available
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(f"\n=== Parallel Processing Mode: {self.parallel_screeners} threads ===")
        
        # 查找所有支持的文件
        file_patterns = [
            os.path.join(folder_path, "*.pdf"),
            os.path.join(folder_path, "*.docx"),
            os.path.join(folder_path, "*.doc"),
            os.path.join(folder_path, "*.txt")
        ]
        
        all_files = []
        for pattern in file_patterns:
            all_files.extend(glob.glob(pattern))
        
        if not all_files:
            print(get_message("system_error", error=f"No supported files found in {folder_path}"))
            return
        
        all_files.sort()
        total_files = len(all_files)
        print(get_message("detected_documents", count=total_files))
        
        # 创建筛选提示词
        if self.combined_prompt is None:
            print(get_message("prompt_generation"))
            self.combined_prompt = self.create_combined_prompt(inclusion_criteria, exclusion_criteria, prompt_file_path)
            print(get_message("prompt_loaded"))
        
        results = []
        start_time = time.time()
        
        try:
            # 使用线程池进行并行处理
            with ThreadPoolExecutor(max_workers=self.parallel_screeners) as executor:
                print(f"\n{get_message('batch_processing')}")
                print(get_message("batch_started", batch_id="Parallel", count=total_files))
                
                # 提交所有任务
                future_to_file = {}
                for i, file_path in enumerate(all_files):
                    # 在筛选器之间添加小延迟，避免 API 限制
                    if i > 0 and i % self.parallel_screeners == 0:
                        time.sleep(self.delay_between_screeners)
                    
                    future = executor.submit(self.process_single_document, file_path, inclusion_criteria, i, total_files)
                    future_to_file[future] = file_path
                
                # 收集结果
                completed = 0
                last_progress_time = time.time()
                
                for future in as_completed(future_to_file):
                    file_path = future_to_file[future]
                    file_name = os.path.basename(file_path)
                    
                    try:
                        file_results = future.result()
                        if file_results:
                            results.extend(file_results)
                            print(get_message("document_completed", filename=file_name))
                        else:
                            print(get_message("document_failed", filename=file_name, error="processing failed"))
                        
                        completed += 1
                        
                        # 显示进度统计（每5秒或每10个文档更新一次）
                        current_time = time.time()
                        if (current_time - last_progress_time > 5) or (completed % 10 == 0) or (completed == total_files):
                            elapsed_time = current_time - start_time
                            speed = completed / (elapsed_time / 60) if elapsed_time > 0 else 0
                            remaining_docs = total_files - completed
                            
                            print(f"\n📊 {get_message('batch_progress', batch_id='Parallel', processed=completed, total=total_files)}")
                            if speed > 0:
                                print(get_message("processing_speed", speed=f"{speed:.1f}"))
                                if remaining_docs > 0:
                                    estimated_remaining = (remaining_docs / speed) * 60  # seconds
                                    remaining_minutes = int(estimated_remaining / 60)
                                    remaining_seconds = int(estimated_remaining % 60)
                                    print(get_message("estimated_time", time=f"{remaining_minutes}m {remaining_seconds}s"))
                            
                            last_progress_time = current_time
                        
                    except Exception as e:
                        print(get_message("document_failed", filename=file_name, error=str(e)))
                        completed += 1
                
                print(f"\n{get_message('all_batches_completed')}")
        
        except KeyboardInterrupt:
            print(f"\n{get_message('task_interrupted')}")
            print("Saving current results...")
        
        # 保存结果到Excel
        self.save_results_to_excel(results, output_path)
        self.save_tokens_to_csv(output_path)
        
        print(f"\n{get_message('parallel_processing_complete')}")
        print(get_message('processed_results_count', count=len(results)))
        
        return output_path
    
    def process_documents(self, folder_path, inclusion_criteria, output_path, exclusion_criteria=None, prompt_file_path=None):
        """处理所有文档 - 根据配置选择单线程或并行模式"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            # Fallback if i18n not available
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(f"\n{get_message('document_processing_mode', mode=self.screening_mode)}")
        
        if self.screening_mode == 'parallel' and self.parallel_screeners > 1:
            # 使用并行处理
            return self.process_documents_parallel(folder_path, inclusion_criteria, output_path, exclusion_criteria, prompt_file_path)
        else:
            # 使用单线程处理
            return self.process_documents_sequential(folder_path, inclusion_criteria, output_path, exclusion_criteria, prompt_file_path)
    
    def process_documents_sequential(self, folder_path, inclusion_criteria, output_path, exclusion_criteria=None, prompt_file_path=None):
        """单线程顺序处理所有文档"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            # Fallback if i18n not available
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(f"\n{get_message('sequential_processing_mode')}")
        
        # 查找所有支持的文件
        file_patterns = [
            os.path.join(folder_path, "*.pdf"),
            os.path.join(folder_path, "*.docx"),
            os.path.join(folder_path, "*.doc"),
            os.path.join(folder_path, "*.txt")
        ]
        
        all_files = []
        for pattern in file_patterns:
            all_files.extend(glob.glob(pattern))
        
        if not all_files:
            print(get_message("system_error", error=f"No supported files found in {folder_path}"))
            return
        
        all_files.sort()
        total_files = len(all_files)
        print(get_message("detected_documents", count=total_files))
        
        # 创建筛选提示词
        if self.combined_prompt is None:
            print(get_message("prompt_generation"))
            self.combined_prompt = self.create_combined_prompt(inclusion_criteria, exclusion_criteria, prompt_file_path)
            print(get_message("prompt_loaded"))
        
        results = []
        start_time = time.time()
        
        try:
            for i, file_path in enumerate(all_files):
                file_name = os.path.basename(file_path)
                print(f"\n{get_message('current_document', filename=file_name)}")
                
                # 获取文件大小
                try:
                    file_size_kb = os.path.getsize(file_path) / 1024
                    print(get_message("document_size", size=f"{file_size_kb:.1f}"))
                except:
                    pass
                
                # 读取文档内容
                print(get_message("pdf_extraction_started"))
                document_content = DocumentReader.read_file(file_path)
                if document_content is None:
                    print(get_message("document_failed", filename=file_name, error="unable to read file"))
                    continue
                
                # 显示提取的文本信息
                char_count = len(document_content)
                print(get_message("text_extracted", chars=f"{char_count:,}"))
                
                # 提取研究ID
                # Get first valid LLM name (skip comment fields)
                valid_llm_names = [name for name in self.screening_llm_configs.keys() if not name.startswith('_')]
                llm_name = valid_llm_names[0] if valid_llm_names else list(self.screening_llm_configs.keys())[0]
                study_id = self.extract_study_id(document_content, llm_name)
                
                # 处理文档（无预筛选）
                doc_start_time = time.time()
                print(get_message("llm_screening_started"))
                study_results = self.process_study_sequential(document_content, inclusion_criteria, study_id)
                doc_end_time = time.time()
                
                print(get_message("document_completed", filename=file_name))
                
                # 显示进度统计
                elapsed_time = doc_end_time - start_time
                avg_time_per_doc = elapsed_time / (i + 1)
                remaining_docs = total_files - (i + 1)
                estimated_remaining = avg_time_per_doc * remaining_docs
                
                print(get_message("documents_processed", processed=i+1, total=total_files))
                if i > 0:  # Only show speed after first document
                    speed = (i + 1) / (elapsed_time / 60)  # docs per minute
                    print(get_message("processing_speed", speed=f"{speed:.1f}"))
                    if remaining_docs > 0:
                        remaining_minutes = int(estimated_remaining / 60)
                        remaining_seconds = int(estimated_remaining % 60)
                        print(get_message("estimated_time", time=f"{remaining_minutes}m {remaining_seconds}s"))
                
                # 保存结果
                for llm_name, result in study_results.items():
                    result_data = {
                        'filename': file_name,
                        'filepath': file_path,
                        'study_id': study_id,
                        'llm_name': llm_name,
                        'picos': result['picos'],
                        'inclusion': result['inclusion']
                    }
                    results.append(result_data)
        
        except KeyboardInterrupt:
            print(f"\n{get_message('task_interrupted')}")
            print("Saving current results...")
        
        # 保存结果到Excel
        print(f"\n{get_message('results_merging')}")
        self.save_results_to_excel(results, output_path)
        self.save_tokens_to_csv(output_path)
        
        # 显示最终统计
        total_time = time.time() - start_time
        print(f"\n{get_message('final_statistics')}")
        print(get_message('total_documents_processed', count=len(results) // len(self.screening_llm_configs)))
        print(get_message('total_processing_time', time=f"{int(total_time // 60)}m {int(total_time % 60)}s"))
        print(get_message('average_time_per_document', time=f"{total_time / max(1, len(results) // len(self.screening_llm_configs)):.1f}"))
        
        return output_path
    
    def process_fulltext_with_prompt(self, llm_name, document_content, prompt, study_id):
        """Process single full-text document with specified prompt"""
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            # Fallback if i18n not available
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        client = self.llm_clients[llm_name]
        config = self.screening_llm_configs[llm_name]
        
        # 截断过长的内容
        estimated_tokens = len(document_content) // 4
        max_tokens = 28000  # 适用于GPT-4和主流大模型
        
        if estimated_tokens > max_tokens:
            print(f"  ⚠️ Content too long (~{estimated_tokens:,} tokens), truncating to ~{max_tokens:,} tokens")
            # 保留开头的75%和结尾的25%内容
            beginning_chars = int(max_tokens * 3)  # 前75%
            ending_chars = int(max_tokens)  # 后25%
            
            content_parts = []
            content_parts.append(document_content[:beginning_chars])
            content_parts.append("\n\n[...CONTENT TRUNCATED FOR LENGTH...]\n\n")
            content_parts.append(document_content[-ending_chars:])
            
            document_content = "".join(content_parts)
        
        print(get_message("sending_to_llm", llm_name=llm_name))
        
        max_retries = 10
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                response = client.chat.completions.create(
                    model=config['model'],
                    messages=[
                        {"role": "system", "content": prompt},
                        {"role": "user", "content": document_content}
                    ],
                    temperature=0.0,
                    max_tokens=3000,
                    timeout=240
                )
                
                print(get_message("llm_response_received", llm_name=llm_name))
                result_text = response.choices[0].message.content
                
                # 记录token消耗
                try:
                    prompt_tokens = response.usage.prompt_tokens
                    completion_tokens = response.usage.completion_tokens
                    total_tokens = response.usage.total_tokens
                    
                    self.tokens_log.append({
                        'timestamp': datetime.now().isoformat(),
                        'operation': 'fulltext_screening',
                        'llm_name': llm_name,
                        'model': config['model'],
                        'prompt_tokens': prompt_tokens,
                        'completion_tokens': completion_tokens,
                        'total_tokens': total_tokens,
                        'study_id': study_id,
                        'document_length': len(document_content)
                    })
                except AttributeError:
                    estimated_prompt_tokens = len(prompt + document_content) // 4
                    estimated_completion_tokens = len(result_text) // 4
                    estimated_total_tokens = estimated_prompt_tokens + estimated_completion_tokens
                    
                    self.tokens_log.append({
                        'timestamp': datetime.now().isoformat(),
                        'operation': 'fulltext_screening',
                        'llm_name': llm_name,
                        'model': config['model'],
                        'prompt_tokens': estimated_prompt_tokens,
                        'completion_tokens': estimated_completion_tokens,
                        'total_tokens': estimated_total_tokens,
                        'study_id': study_id,
                        'document_length': len(document_content)
                    })
                
                # 集成解析逻辑：使用=== split和逐行解析（兼容标题摘要Prompt）
                picos_section = ""
                decision_section = ""
                
                if "===PICOS EXTRACTION===" in result_text and "===INCLUSION ASSESSMENT===" in result_text:
                    sections = result_text.split("===INCLUSION ASSESSMENT===")
                    picos_section = sections[0].replace("===PICOS EXTRACTION===", "").strip()
                    decision_section = sections[1].strip()
                else:
                    picos_section = result_text
                    if "Decision:" in result_text:
                        decision_idx = result_text.find("Decision:")
                        decision_section = result_text[decision_idx:].strip()
                
                # Parse PICOS information
                picos = {}
                all_not_reported = True
                
                for line in picos_section.split('\n'):
                    line = line.strip()
                    if not line:
                        continue
                    
                    if line.startswith("Study design:"):
                        picos["study_design"] = line.split(":", 1)[1].strip()
                        if "not reported" not in picos["study_design"].lower():
                            all_not_reported = False
                    elif line.startswith("Participants:"):
                        picos["participants"] = line.split(":", 1)[1].strip()
                        if "not reported" not in picos["participants"].lower():
                            all_not_reported = False
                    elif line.startswith("Intervention:"):
                        picos["intervention"] = line.split(":", 1)[1].strip()
                        if "not reported" not in picos["intervention"].lower():
                            all_not_reported = False
                    elif line.startswith("Comparison:"):
                        picos["comparison"] = line.split(":", 1)[1].strip()
                        if "not reported" not in picos["comparison"].lower():
                            all_not_reported = False
                    elif line.startswith("Outcomes:"):
                        picos["outcomes"] = line.split(":", 1)[1].strip()
                        if "not reported" not in picos["outcomes"].lower():
                            all_not_reported = False
                
                # Parse inclusion decision and explanation
                inclusion_result = "⭕️ UNCLEAR"
                explanation = ""
                short_reason = ""
                
                if "Explanation:" in decision_section:
                    parts = decision_section.split("Explanation:", 1)
                    explanation = parts[1].strip()
                    short_reason = explanation[:300]
                    if "." in short_reason:
                        short_reason = short_reason.split(".", 1)[0] + "."
                
                if "Decision:" in decision_section:
                    decision_line = decision_section.split("\n")[0]
                    decision_value = decision_line.split("Decision:", 1)[1].strip()
                else:
                    decision_value = "UNCLEAR"
                    
                if all_not_reported:
                    inclusion_result = "⭕️ UNCLEAR (insufficient information)"
                else:
                    if "INCLUDE" in decision_value.upper():
                        inclusion_result = "✅ INCLUDE"
                    elif "EXCLUDE" in decision_value.upper():
                        inclusion_result = f"❌ EXCLUDE - {short_reason}"
                    else:
                        inclusion_result = f"⭕️ UNCLEAR - {short_reason}"
                
                return picos, inclusion_result
                
            except Exception as e:
                if "timeout" in str(e).lower() or "timed out" in str(e).lower():
                    retry_count += 1
                    print(f"Timeout error with {llm_name}, attempt {retry_count}/{max_retries}: Response time exceeded 4 minutes. Retrying...")
                    if retry_count >= max_retries:
                        print(f"Failed after {max_retries} timeout attempts with {llm_name}")
                        return self._create_error_result(f"Response timeout after multiple attempts")
                else:
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"Error with {llm_name}, attempt {retry_count}/{max_retries}: {str(e)}. Retrying...")
                        time.sleep(2 ** retry_count)
                    else:
                        print(f"Failed after {max_retries} attempts with {llm_name}: {str(e)}")
                        return self._create_error_result(f"Error in processing after multiple retries: {str(e)}")

    def _create_error_result(self, error_msg):
        """Create error result structure"""
        error_picos = {
            'study_design': f"Error: {error_msg}",
            'participants': f"Error: {error_msg}",
            'intervention': f"Error: {error_msg}",
            'comparison': f"Error: {error_msg}",
            'outcomes': f"Error: {error_msg}"
        }
        error_inclusion = f"⭕ UNCLEAR (Error: {error_msg})"
        return error_picos, error_inclusion

    def save_results_to_excel(self, results, output_path):
        """Save results to Excel file with one row per study"""
        if not results:
            print("No results to save")
            return
        
        # 按文件名分组结果
        grouped_results = {}
        for result in results:
            filename = result['filename']
            if filename not in grouped_results:
                grouped_results[filename] = {
                    'filename': filename,
                    'study_id': result['study_id'],
                    'llm_results': {}
                }
            grouped_results[filename]['llm_results'][result['llm_name']] = result
        
        # 创建DataFrame
        data = []
        # Get valid LLM names (filter out comment fields)
        all_llm_names = list(self.screening_llm_configs.keys())
        llm_names = [name for name in all_llm_names if not name.startswith('_')]
        
        for filename, group_data in grouped_results.items():
            row = {
                'Filename': group_data['filename'],
                'Study ID': group_data['study_id'],
                'Last Update': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # 为每个LLM添加列
            for llm_name in llm_names:
                if llm_name in group_data['llm_results']:
                    llm_result = group_data['llm_results'][llm_name]
                    row[f'Decision ({llm_name})'] = llm_result['inclusion']
                    row[f'Study Design ({llm_name})'] = llm_result['picos'].get('study_design', '')
                    row[f'Participants ({llm_name})'] = llm_result['picos'].get('participants', '')
                    row[f'Intervention ({llm_name})'] = llm_result['picos'].get('intervention', '')
                    row[f'Comparison ({llm_name})'] = llm_result['picos'].get('comparison', '')
                    row[f'Outcomes ({llm_name})'] = llm_result['picos'].get('outcomes', '')
                else:
                    # 如果某个LLM没有结果，填入空值
                    row[f'Decision ({llm_name})'] = 'Not processed'
                    row[f'Study Design ({llm_name})'] = ''
                    row[f'Participants ({llm_name})'] = ''
                    row[f'Intervention ({llm_name})'] = ''
                    row[f'Comparison ({llm_name})'] = ''
                    row[f'Outcomes ({llm_name})'] = ''
            
            # 添加冲突状态
            row['Conflict Status'] = self._determine_conflict_status(group_data['llm_results'])
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # 保存到Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Full Text Results', index=False)
            
            # 获取工作表进行格式化
            worksheet = writer.sheets['Full Text Results']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 25  # Filename
            worksheet.column_dimensions['B'].width = 15  # Study ID
            worksheet.column_dimensions['C'].width = 18  # Last Update
            
            # 为每个LLM的列设置宽度
            col_index = 4  # 从第4列开始
            for llm_name in llm_names:
                worksheet.column_dimensions[chr(64 + col_index)].width = 30      # Decision
                worksheet.column_dimensions[chr(64 + col_index + 1)].width = 20  # Study Design
                worksheet.column_dimensions[chr(64 + col_index + 2)].width = 25  # Participants
                worksheet.column_dimensions[chr(64 + col_index + 3)].width = 25  # Intervention
                worksheet.column_dimensions[chr(64 + col_index + 4)].width = 20  # Comparison
                worksheet.column_dimensions[chr(64 + col_index + 5)].width = 25  # Outcomes
                col_index += 6
            
            worksheet.column_dimensions[chr(64 + col_index)].width = 15  # Conflict Status
            
            # 设置标题行格式
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            # 设置数据行格式和颜色
            for row in range(2, len(df) + 2):
                # 获取冲突状态
                conflict_status = worksheet.cell(row=row, column=len(df.columns)).value
                
                # 设置行颜色
                row_fill = None
                if conflict_status == "Conflict":
                    row_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 红色 - 冲突
                elif conflict_status == "Uncertain":
                    row_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 黄色 - 不确定
                elif conflict_status == "Include":
                    row_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 绿色 - 一致纳入
                elif conflict_status == "Exclude":
                    row_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # 灰色 - 一致排除
                
                # 应用行格式
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if row_fill:
                        cell.fill = row_fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # 冻结顶部行和前3列
            worksheet.freeze_panes = "D2"
        
        # Import i18n functions
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            # Fallback if i18n not available
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(get_message("results_saved", path=output_path))
        
        # 显示筛选统计
        self._display_screening_statistics(grouped_results)
        
        # 显示Token使用统计
        self._display_token_statistics()

    def _display_screening_statistics(self, grouped_results):
        """Display screening statistics"""
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        print(f"\n📊 {get_message('screening_criteria_loaded')}")
        
        # 统计各种决策结果
        include_count = 0
        exclude_count = 0
        unclear_count = 0
        
        for filename, group_data in grouped_results.items():
            # 获取第一个LLM的决策作为主要决策
            llm_results = group_data['llm_results']
            if llm_results:
                first_result = next(iter(llm_results.values()))
                decision = first_result['inclusion']
                
                if "✅ INCLUDE" in decision or "INCLUDE" in decision.upper():
                    include_count += 1
                elif "❌ EXCLUDE" in decision or "EXCLUDE" in decision.upper():
                    exclude_count += 1
                else:
                    unclear_count += 1
        
        print(get_message("documents_included", count=include_count))
        print(get_message("documents_excluded", count=exclude_count))
        print(get_message("documents_unclear", count=unclear_count))
        
        total_docs = include_count + exclude_count + unclear_count
        if total_docs > 0:
            print(f"  • Inclusion rate: {include_count/total_docs*100:.1f}%")
            print(f"  • Exclusion rate: {exclude_count/total_docs*100:.1f}%")
            if unclear_count > 0:
                print(f"  • Unclear rate: {unclear_count/total_docs*100:.1f}%")
    
    def _display_token_statistics(self):
        """Display token usage statistics"""
        try:
            from i18n.i18n_manager import get_message
        except ImportError:
            def get_message(key, **kwargs):
                return key.format(**kwargs) if kwargs else key
        
        if not self.tokens_log:
            return
        
        print(f"\n💰 {get_message('token_usage')}")
        
        total_tokens = sum(entry['total_tokens'] for entry in self.tokens_log)
        total_prompt_tokens = sum(entry['prompt_tokens'] for entry in self.tokens_log)
        total_completion_tokens = sum(entry['completion_tokens'] for entry in self.tokens_log)
        
        print(get_message("total_tokens", total=f"{total_tokens:,}"))
        print(f"  • Prompt tokens: {total_prompt_tokens:,}")
        print(f"  • Completion tokens: {total_completion_tokens:,}")
        
        # 计算平均每文档Token使用
        unique_studies = len(set(entry['study_id'] for entry in self.tokens_log))
        if unique_studies > 0:
            avg_tokens_per_doc = total_tokens / unique_studies
            print(get_message("average_tokens_per_doc", avg=f"{avg_tokens_per_doc:,.0f}"))
        
        # 估算成本（基于常见模型定价）
        estimated_cost = self._estimate_cost(total_prompt_tokens, total_completion_tokens)
        if estimated_cost > 0:
            print(get_message("tokens_used", tokens=f"{total_tokens:,}", cost=estimated_cost))
    
    def _estimate_cost(self, prompt_tokens, completion_tokens):
        """Estimate cost based on token usage"""
        # 简单的成本估算，基于常见模型定价
        # 这里使用GPT-4的定价作为参考：$0.03/1K prompt tokens, $0.06/1K completion tokens
        prompt_cost = (prompt_tokens / 1000) * 0.03
        completion_cost = (completion_tokens / 1000) * 0.06
        return prompt_cost + completion_cost

    def _determine_conflict_status(self, llm_results):
        """Determine conflict status based on LLM decisions"""
        decisions = []
        for llm_name, result in llm_results.items():
            inclusion = result['inclusion']
            if "✅ INCLUDE" in inclusion:
                decisions.append("INCLUDE")
            elif "❌ EXCLUDE" in inclusion:
                decisions.append("EXCLUDE")
            else:
                decisions.append("UNCLEAR")
        
        # 判断冲突状态
        unique_decisions = set(decisions)
        
        if "UNCLEAR" in unique_decisions:
            return "Uncertain"  # 至少一个LLM不确定
        elif len(unique_decisions) > 1:
            return "Conflict"   # LLM之间有分歧
        elif "INCLUDE" in unique_decisions:
            return "Include"    # 一致纳入
        elif "EXCLUDE" in unique_decisions:
            return "Exclude"    # 一致排除
        else:
            return "Unknown"


    def save_tokens_to_csv(self, output_path):
        """Save token consumption statistics to CSV file"""
        if not self.tokens_log:
            print("No token consumption records")
            return
        
        # 生成CSV文件路径
        if self.tokens_csv_path is None:
            base_name = os.path.splitext(output_path)[0]
            self.tokens_csv_path = f"{base_name}_tokens_usage.csv"
        
        # 写入CSV文件
        fieldnames = ['timestamp', 'operation', 'llm_name', 'model', 'prompt_tokens', 
                    'completion_tokens', 'total_tokens', 'study_id', 'document_length']
        
        with open(self.tokens_csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.tokens_log)
        
        # 计算总计
        total_prompt_tokens = sum(log['prompt_tokens'] for log in self.tokens_log)
        total_completion_tokens = sum(log['completion_tokens'] for log in self.tokens_log)
        total_tokens = sum(log['total_tokens'] for log in self.tokens_log)
        
        print(f"\n=== TOKEN CONSUMPTION STATISTICS ===")
        print(f"Detailed records saved to: {self.tokens_csv_path}")
        print(f"Total: {total_tokens:,} tokens (Input: {total_prompt_tokens:,}, Output: {total_completion_tokens:,})")
